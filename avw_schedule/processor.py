from __future__ import annotations

from dataclasses import dataclass, asdict
from io import BytesIO
import math
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCHEDULE_COLUMNS = [
    "Order",
    "Group ID",
    "Parent Group ID",
    "Row ID",
    "Parent Row ID",
    "Depth",
    "Instance",
    "Nested Code",
    "Is Assembly Root",
    "Item",
    "Description",
    "Count Category",
    "#",
    "Qty",
    "Source Quote Qty",
    "HP",
    "Phase",
    "Volts",
    "Amps",
    "Ext. Amps",
    "C.B.",
    "Air Port",
    "Cold Water",
    "Hot Water",
    "Reclaim Water",
    "Gas BTUH",
    "Match Type",
    "Source Quote Item",
    "Master Excel Row",
    "Requirements Source",
    "Requirements Status",
    "Bold",
    "Italic",
    "Underline",
    "Highlight",
    "Review Added",
]

REQUIREMENT_COLUMNS = [
    "HP",
    "Phase",
    "Volts",
    "Amps",
    "C.B.",
    "Air Port",
    "Cold Water",
    "Hot Water",
    "Reclaim Water",
    "Gas BTUH",
]

QUOTE_COLUMNS = [
    "Order",
    "Item",
    "Description",
    "#",
    "Qty",
    "Unit Price",
    "Total Price",
    "Page",
    "Notes",
]

REVIEW_COLUMNS = [
    "Order",
    "Item",
    "Description",
    "#",
    "Qty",
    "Issue Type",
    "Details",
    "Candidate Lookup",
]

# Engineering-approved canonical code changes. These aliases keep older quote
# PDFs usable after the Master List is renamed; they do not perform fuzzy or
# description-based matching.
CANONICAL_CODE_ALIASES = {
    "CN1": "CN1-3524",
    "WA1-SK": "WA1-SK-2018",
    "WA1-SKP": "WA1-SKP-2018",
    "WA1P": "WA1-120V",
}
WRAP_AIR_ASSIST_KIT = "WA1-SK-2018"
WRAP_AIR_ASSIST_PANEL = "WA1-SKP-2018"


def norm(value: Any) -> str:
    """Normalize part numbers for deterministic matching."""
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    text = text.replace("…", "...")
    text = re.sub(r"\s+", "", text)
    return text.upper()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def enabled_bool(value: Any) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y", "enabled"}


def safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text or text == "-":
        return None
    text = text.replace(",", "").replace("T", "").replace("$", "")
    try:
        return float(text)
    except ValueError:
        return None


def is_whole_number(value: Any) -> bool:
    number = safe_float(value)
    if number is None:
        return False
    return abs(number - round(number)) < 1e-9


def qty_to_instances(qty: Any) -> int:
    number = safe_float(qty)
    if number is None:
        return 1
    if abs(number - round(number)) < 1e-9 and number >= 1:
        return int(round(number))
    return 1


def normalize_voltage(value: Any) -> str:
    text = clean_text(value)
    if not text or text == "-":
        return ""
    text = text.upper().replace("VAC", "V").strip()
    text = re.sub(r"\s+", "", text)
    if re.fullmatch(r"\d+(\.0+)?", text):
        text = str(int(float(text))) + "V"
    elif re.fullmatch(r"\d+(\.0+)?V", text):
        text = str(int(float(text[:-1]))) + "V"
    return text


def amp_value(value: Any) -> float:
    number = safe_float(value)
    return 0.0 if number is None else number


def extended_amp_value(amps: Any, qty: Any) -> float:
    """Calculate row total amps using quote quantity without duplicating rows."""
    amp = amp_value(amps)
    quantity = safe_float(qty)
    if quantity is None:
        quantity = 1.0
    return round(amp * quantity, 4)


def component_count_category(item: Any, description: Any) -> str:
    """Return the approved counting bucket used by the schedule # column.

    Motor SKUs have different ratings but share one Motor sequence. Other
    components default to their exact canonical part number until engineering
    assigns another category in the maintained resolution library.
    """
    item_key = norm(item)
    desc = clean_text(description).upper().lstrip("- ")
    if item_key == "M" or item_key.startswith("M-") or desc == "MOTOR":
        return "Motor"
    return item_key


@dataclass
class QuoteLine:
    order: int
    item: str
    description: str
    qty: Any
    unit_price: Any = ""
    total_price: Any = ""
    page: int = 0
    notes: str = ""


@dataclass
class MasterRow:
    excel_row: int
    group_id: int
    nested_code: str
    part_number: str
    description: str
    hp: Any
    phase: Any
    volts: Any
    amps: Any
    cb: Any
    air_port: Any
    cold_water: Any
    hot_water: Any
    reclaim_water: Any
    gas_btuh: Any
    is_parent: bool
    is_bold: bool = False

    def to_schedule_row(
        self,
        order: str,
        group_id: str,
        qty: Any,
        quote_item: str,
        quote_description: str,
        match_type: str,
        instance: int = 1,
        source_quote_qty: Any = 1,
    ) -> Dict[str, Any]:
        description = quote_description if self.is_parent and quote_description else self.description
        return {
            "Order": order,
            "Group ID": group_id,
            "Parent Group ID": "",
            "Row ID": "",
            "Parent Row ID": "",
            "Depth": 0 if self.is_parent else max(1, len(clean_text(self.nested_code))),
            "Instance": instance,
            "Nested Code": "Parent" if self.is_parent else self.nested_code,
            "Is Assembly Root": bool(self.is_parent),
            "Item": self.part_number,
            "Description": description,
            "Count Category": component_count_category(self.part_number, description),
            "#": "",
            "Qty": qty,
            "Source Quote Qty": source_quote_qty,
            "HP": self.hp,
            "Phase": self.phase,
            "Volts": self.volts,
            "Amps": self.amps,
            "Ext. Amps": extended_amp_value(self.amps, qty),
            "C.B.": self.cb,
            "Air Port": self.air_port,
            "Cold Water": self.cold_water,
            "Hot Water": self.hot_water,
            "Reclaim Water": self.reclaim_water,
            "Gas BTUH": self.gas_btuh,
            "Match Type": match_type,
            "Source Quote Item": quote_item,
            "Master Excel Row": self.excel_row,
            "Requirements Source": f"Master row {self.excel_row} · {self.part_number}",
            "Requirements Status": "Inherited from Master",
            "Bold": bool(self.is_parent or self.is_bold),
            "Italic": False,
            "Underline": False,
            "Highlight": False,
            "Review Added": False,
        }


@dataclass
class QuickBooksItem:
    excel_row: int
    item_number: str
    description: str
    primary_category: str = ""
    schedule_status: str = ""
    active_status: str = ""
    quickbooks_category: str = ""

    def to_schedule_row(
        self,
        order: str,
        group_id: str,
        qty: Any,
        quote_item: str,
        match_type: str,
        instance: int = 1,
        source_quote_qty: Any = 1,
    ) -> Dict[str, Any]:
        """Map a QuickBooks item to one standalone schedule row.

        QuickBooks is an item/description fallback source only. Engineering
        requirements are deliberately left blank and visibly flagged so the
        user never mistakes accounting data for Master List requirements.
        """
        return {
            "Order": order,
            "Group ID": group_id,
            "Parent Group ID": "",
            "Row ID": "",
            "Parent Row ID": "",
            "Depth": 0,
            "Instance": instance,
            "Nested Code": "Parent",
            "Is Assembly Root": False,
            "Item": self.item_number,
            "Description": self.description,
            "Count Category": component_count_category(self.item_number, self.description),
            "#": "",
            "Qty": qty,
            "Source Quote Qty": source_quote_qty,
            "HP": "-",
            "Phase": "-",
            "Volts": "-",
            "Amps": "-",
            "Ext. Amps": 0,
            "C.B.": "-",
            "Air Port": "-",
            "Cold Water": "-",
            "Hot Water": "-",
            "Reclaim Water": "-",
            "Gas BTUH": "-",
            "Match Type": match_type,
            "Source Quote Item": quote_item,
            "Master Excel Row": "",
            "Requirements Source": (
                f"QuickBooks Unique Items row {self.excel_row} · {self.item_number}"
            ),
            "Requirements Status": "QuickBooks fallback · engineering requirements needed",
            "Bold": False,
            "Italic": False,
            "Underline": False,
            "Highlight": False,
            "Review Added": False,
        }


class QuickBooksCatalog:
    """Indexed QuickBooks item/description fallback catalog."""

    def __init__(self, items: List[QuickBooksItem]):
        self.items = items
        self.item_index: Dict[str, List[QuickBooksItem]] = {}
        self.row_index: Dict[int, QuickBooksItem] = {}
        for item in items:
            key = norm(item.item_number)
            if key:
                self.item_index.setdefault(key, []).append(item)
            self.row_index[item.excel_row] = item

    @classmethod
    def from_excel(
        cls,
        file_obj: Any,
        sheet_name: str = "Unique Items",
    ) -> "QuickBooksCatalog":
        wb = load_workbook(file_obj, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

        header_row = 0
        header_map: Dict[str, int] = {}
        item_rows: List[Tuple[int, Tuple[Any, ...]]] = []
        for row_number, raw_values in enumerate(ws.iter_rows(values_only=True), start=1):
            if not header_row:
                if row_number > 30:
                    break
                values = [clean_text(cell) for cell in raw_values]
                normalized = {
                    value.lower(): index for index, value in enumerate(values) if value
                }
                if "item number" in normalized and "description" in normalized:
                    header_row = row_number
                    header_map = normalized
                continue
            item_rows.append((row_number, tuple(raw_values)))
        if not header_row:
            raise ValueError(
                "QuickBooks workbook must contain Item Number and Description headers."
            )

        def value(raw_values: Tuple[Any, ...], header: str) -> str:
            column = header_map.get(header.lower())
            if column is None or column >= len(raw_values):
                return ""
            return clean_text(raw_values[column])

        items: List[QuickBooksItem] = []
        for row_number, raw_values in item_rows:
            item_number = value(raw_values, "Item Number")
            description = value(raw_values, "Description")
            if not item_number:
                continue
            items.append(QuickBooksItem(
                excel_row=row_number,
                item_number=item_number,
                description=description,
                primary_category=value(raw_values, "Primary Category"),
                schedule_status=value(raw_values, "9-Schedule Status"),
                active_status=value(raw_values, "Active Status"),
                quickbooks_category=value(raw_values, "QuickBooks Category"),
            ))
        wb.close()
        return cls(items)

    def find_exact(self, item_number: Any) -> Tuple[str, List[QuickBooksItem]]:
        matches = self.item_index.get(norm(item_number), [])
        if not matches:
            return "none", []
        if len(matches) == 1:
            return "single", matches
        signatures = {
            (norm(item.item_number), clean_text(item.description).upper())
            for item in matches
        }
        if len(signatures) == 1:
            return "identical_duplicate", matches
        return "ambiguous", matches


class MasterList:
    def __init__(self, rows: List[MasterRow], main_equipment_reference: List[str]):
        self.rows = rows
        self.main_equipment_reference = main_equipment_reference
        self.groups: Dict[int, List[MasterRow]] = {}
        self.parent_index: Dict[str, List[MasterRow]] = {}
        self.component_index: Dict[str, List[MasterRow]] = {}
        for row in rows:
            self.groups.setdefault(row.group_id, []).append(row)
            if row.part_number:
                self.component_index.setdefault(norm(row.part_number), []).append(row)
                if row.is_parent:
                    self.parent_index.setdefault(norm(row.part_number), []).append(row)

    @classmethod
    def from_excel(cls, file_obj: Any, sheet_name: str = "MASTER LIST") -> "MasterList":
        wb = load_workbook(file_obj, data_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

        rows: List[MasterRow] = []
        main_ref: List[str] = []
        current_group_id = 0
        found_data = False

        # Known AVW layout: data starts near row 7, columns B-Q carry schedule data, AF is helper only.
        for r in range(1, ws.max_row + 1):
            part = clean_text(ws.cell(r, 4).value)  # D
            desc = clean_text(ws.cell(r, 6).value)  # F
            nested = clean_text(ws.cell(r, 3).value)  # C
            side_ref = clean_text(ws.cell(r, 32).value)  # AF
            if side_ref and side_ref.upper() not in {"MAIN EQUIPMENT LIST", "NONE"}:
                main_ref.append(side_ref)

            # Start reading once likely schedule rows begin.
            if r < 7:
                continue
            # A component description can legitimately begin with "TOTAL"
            # (for example, "TOTAL DISSOLVED SOLIDS MONITOR").  Treat it as
            # the schedule footer only when the row has no part or nested code.
            is_description_footer = bool(
                not part and not nested and desc.upper().startswith("TOTAL")
            )
            if is_description_footer or part.upper() in {"TOTAL", "SUBTOTAL"}:
                break
            if not part and not desc:
                continue

            # AVW parent rows are bold rows with a project item number and part
            # number. Nested code remains a defensive guard against bold child
            # formatting accidentally creating a new group.
            project_item = clean_text(ws.cell(r, 2).value)
            is_bold = bool(ws.cell(r, 4).font.bold or ws.cell(r, 6).font.bold)
            is_parent = bool(part and project_item and is_bold and not nested)
            if is_parent:
                current_group_id += 1
            elif current_group_id == 0:
                # Defensive fallback: child before a parent gets its own group.
                current_group_id += 1

            rows.append(
                MasterRow(
                    excel_row=r,
                    group_id=current_group_id,
                    nested_code=nested,
                    part_number=part,
                    description=desc,
                    hp=ws.cell(r, 7).value if ws.cell(r, 7).value is not None else "-",
                    phase=ws.cell(r, 8).value if ws.cell(r, 8).value is not None else "-",
                    volts=ws.cell(r, 9).value if ws.cell(r, 9).value is not None else "-",
                    amps=ws.cell(r, 10).value if ws.cell(r, 10).value is not None else "-",
                    cb=ws.cell(r, 11).value if ws.cell(r, 11).value is not None else "-",
                    air_port=ws.cell(r, 12).value if ws.cell(r, 12).value is not None else "-",
                    cold_water=ws.cell(r, 13).value if ws.cell(r, 13).value is not None else "-",
                    hot_water=ws.cell(r, 14).value if ws.cell(r, 14).value is not None else "-",
                    reclaim_water=ws.cell(r, 15).value if ws.cell(r, 15).value is not None else "-",
                    gas_btuh=ws.cell(r, 17).value if ws.cell(r, 17).value is not None else "-",
                    is_parent=is_parent,
                    is_bold=is_bold,
                )
            )
            found_data = True

        return cls(rows, main_ref)

    def group_for_parent(self, parent_row: MasterRow) -> List[MasterRow]:
        rows = self.groups.get(parent_row.group_id, [parent_row])
        # PSH25 is an optional pumping station, not an automatic component of
        # PWB1. Remove its complete nested subtree even when an older Master
        # workbook still stores it beneath PWB1.
        if norm(parent_row.part_number) != "PWB1":
            return rows
        blocked_prefixes = [
            clean_text(row.nested_code).upper()
            for row in rows
            if norm(row.part_number) == "PSH25" and clean_text(row.nested_code)
        ]
        if not blocked_prefixes:
            return rows
        return [
            row for row in rows
            if not any(
                clean_text(row.nested_code).upper() == prefix
                or (
                    clean_text(row.nested_code).upper().startswith(prefix)
                    and len(clean_text(row.nested_code)) > len(prefix)
                )
                for prefix in blocked_prefixes
            )
        ]

    def find_parent(self, item: str) -> Tuple[str, List[MasterRow]]:
        matches = self.parent_index.get(norm(item), [])
        if len(matches) == 1:
            return "single", matches
        if len(matches) > 1:
            signatures = {self.group_signature(m.group_id) for m in matches}
            if len(signatures) == 1:
                return "identical_duplicate", [matches[0]]
            return "ambiguous", matches
        return "none", []

    def find_component(self, item: str) -> Tuple[str, List[MasterRow]]:
        matches = self.component_index.get(norm(item), [])
        if len(matches) == 1:
            return "single", matches
        if len(matches) > 1:
            row_sigs = {self.component_signature(m) for m in matches}
            if len(row_sigs) == 1:
                return "identical_duplicate", [matches[0]]
            return "ambiguous", matches
        return "none", []

    def group_signature(self, group_id: int) -> Tuple:
        rows = self.groups.get(group_id, [])
        return tuple(self.row_signature(r) for r in rows)

    @staticmethod
    def row_signature(row: MasterRow) -> Tuple:
        """Strict row signature used for parent group duplicate checks."""
        return (
            norm(row.part_number),
            row.nested_code,
            clean_text(row.description).upper(),
            clean_text(row.hp),
            clean_text(row.phase),
            clean_text(row.volts),
            clean_text(row.amps),
            clean_text(row.cb),
            clean_text(row.air_port),
            clean_text(row.cold_water),
            clean_text(row.hot_water),
            clean_text(row.reclaim_water),
            clean_text(row.gas_btuh),
        )

    @staticmethod
    def component_signature(row: MasterRow) -> Tuple:
        """Engineering-only signature for duplicate component auto-resolve.

        Two component rows with the same part number and same engineering
        requirements are safe to auto-resolve even when the nested code or
        surrounding parent package is different. This fixes cases like
        VA3246S appearing as a child in more than one parent group.
        """
        return (
            norm(row.part_number),
            clean_text(row.hp),
            clean_text(row.phase),
            clean_text(row.volts),
            clean_text(row.amps),
            clean_text(row.cb),
            clean_text(row.air_port),
            clean_text(row.cold_water),
            clean_text(row.hot_water),
            clean_text(row.reclaim_water),
            clean_text(row.gas_btuh),
        )

    def candidate_part_numbers_starting_with(self, prefix: str) -> List[str]:
        """Return unique Master List part numbers that start with a truncated prefix."""
        prefix_norm = norm(prefix)
        if not prefix_norm:
            return []
        out: Dict[str, str] = {}
        for row in self.rows:
            part_norm = norm(row.part_number)
            if part_norm and part_norm.startswith(prefix_norm):
                out.setdefault(part_norm, row.part_number)
        return list(out.values())


def read_rules_csv(uploaded: Any) -> pd.DataFrame:
    if uploaded is None:
        return pd.DataFrame()
    return pd.read_csv(uploaded).fillna("")


def build_alias_map(alias_df: pd.DataFrame) -> Dict[str, str]:
    if alias_df.empty:
        return {}
    out: Dict[str, str] = {}
    for _, row in alias_df.iterrows():
        if "enabled" in row and not enabled_bool(row.get("enabled")):
            continue
        quote_item = norm(row.get("quote_item"))
        lookup = clean_text(row.get("master_lookup_item"))
        if quote_item and lookup:
            out[quote_item] = lookup
    return out


def build_replacement_map(rep_df: pd.DataFrame) -> Dict[str, List[Tuple[str, str]]]:
    if rep_df.empty:
        return {}
    out: Dict[str, List[Tuple[str, str]]] = {}
    for _, row in rep_df.iterrows():
        if "enabled" in row and not enabled_bool(row.get("enabled")):
            continue
        quote_item = norm(row.get("quote_item"))
        replacement_item = clean_text(row.get("replacement_item"))
        master_lookup = clean_text(row.get("master_lookup_item")) or replacement_item
        if quote_item and replacement_item:
            out.setdefault(quote_item, []).append((replacement_item, master_lookup))
    return out


def build_special_map(special_df: pd.DataFrame) -> Dict[str, Dict[str, str]]:
    if special_df.empty:
        return {}
    out: Dict[str, Dict[str, str]] = {}
    for _, row in special_df.iterrows():
        if "enabled" in row and not enabled_bool(row.get("enabled")):
            continue
        item = norm(row.get("item"))
        if item:
            out[item] = {
                "action": clean_text(row.get("action")),
                "note": clean_text(row.get("note")),
            }
    return out


def build_ignore_map(ignore_df: pd.DataFrame) -> Dict[str, str]:
    if ignore_df.empty:
        return {}
    out: Dict[str, str] = {}
    for _, row in ignore_df.iterrows():
        if "enabled" in row and not enabled_bool(row.get("enabled")):
            continue
        item = norm(row.get("item"))
        if item:
            out[item] = clean_text(row.get("reason"))
    return out


_ITEM_RE = re.compile(r"^[A-Z0-9][A-Z0-9_.\-/]*\.?\.?(?:\.\.\.)?$", re.I)
_PRICE_END_RE = re.compile(r"(?P<qty>-?\d+(?:\.\d+)?)\s+(?P<unit>[\d,]+\.\d{2})\s+(?P<total>-?[\d,]+\.\d{2})T?$")
_SKIP_ITEM_WORDS = {
    "ITEM", "SUBTOTAL", "TOTAL", "CUSTOMER", "DEPOSITS", "PAGE", "ACKNOWLEDGEMENT/DEPOSIT", "INVOICE"
}


def is_possible_item_token(text: str) -> bool:
    text = clean_text(text)
    if not text or text.upper() in _SKIP_ITEM_WORDS:
        return False
    if len(text) > 80:
        return False
    if not _ITEM_RE.match(text):
        return False
    # AVW part codes usually contain a digit/hyphen, or are short uppercase codes like RC4/PGS1/UW2/WS1.
    return any(c.isdigit() for c in text) or "-" in text or (len(text) <= 8 and text == text.upper())


def is_qty_line(text: str) -> bool:
    return re.fullmatch(r"-?\d+(?:\.\d+)?", clean_text(text)) is not None


def is_price_line(text: str) -> bool:
    return re.fullmatch(r"-?[\d,]+\.\d{2}T?", clean_text(text)) is not None


def parse_quote_pdf(file_obj: Any) -> Tuple[Dict[str, Any], List[QuoteLine], str]:
    """Extract quote metadata and line items from a text-based AVW / QuickBooks PDF.

    The AVW PDF often stores the table as separate text lines:
    ITEM, description line(s), Qty, Unit Price, Total.
    This parser intentionally avoids OCR and avoids description-based matching.
    """
    data = file_obj.read() if hasattr(file_obj, "read") else file_obj
    doc = fitz.open(stream=data, filetype="pdf")
    full_text = "\n".join(page.get_text("text") for page in doc)
    meta = extract_quote_meta(full_text)

    rows: List[QuoteLine] = []
    order = 0

    for page_index, page in enumerate(doc, start=1):
        lines = [clean_text(x) for x in page.get_text("text").splitlines()]
        lines = [x for x in lines if x]

        # Work only inside the table area. Header columns appear as separate lines.
        try:
            table_start = next(i for i, x in enumerate(lines) if x.upper() == "ITEM")
        except StopIteration:
            table_start = 0
        i = table_start + 1
        while i < len(lines):
            line = lines[i]
            upper = line.upper()
            if upper in {"DESCRIPTION", "QTY", "UNIT PRICE", "TOTAL"}:
                i += 1
                continue
            if upper.startswith("PAGE ") or upper.startswith("SUBTOTAL") or upper.startswith("TOTAL $"):
                break
            initial_desc = ""
            if is_possible_item_token(line):
                item = line
            else:
                pieces = line.split(" ", 1)
                if len(pieces) == 2 and is_possible_item_token(pieces[0]):
                    item = pieces[0]
                    initial_desc = pieces[1]
                else:
                    i += 1
                    continue

            desc_parts: List[str] = [initial_desc] if initial_desc else []
            notes = ""
            qty = unit = total = ""
            j = i + 1
            while j < len(lines):
                cur = lines[j]
                cur_upper = cur.upper()
                if cur_upper in {"ITEM", "DESCRIPTION", "QTY", "UNIT PRICE", "TOTAL"}:
                    j += 1
                    continue
                if cur_upper.startswith("PAGE ") or cur_upper.startswith("SUBTOTAL") or cur_upper.startswith("TOTAL $"):
                    break

                if cur_upper.startswith("REPLACED WITH"):
                    repl = ""
                    if j + 1 < len(lines):
                        repl = lines[j + 1]
                        j += 2
                    else:
                        j += 1
                    notes = f"Replaced with: {repl}".strip()
                    continue

                # Separate-line qty, unit price, total.
                if is_qty_line(cur) and j + 2 < len(lines) and is_price_line(lines[j + 1]) and is_price_line(lines[j + 2]):
                    qty = cur
                    unit = lines[j + 1].replace("T", "")
                    total = lines[j + 2].replace("T", "")
                    j += 3
                    break

                # Defensive support for same-line trailer.
                m = _PRICE_END_RE.search(cur)
                if m:
                    before = cur[:m.start()].strip()
                    if before:
                        desc_parts.append(before)
                    qty, unit, total = m.group("qty"), m.group("unit"), m.group("total")
                    j += 1
                    break

                # Skip repeated headers within the table.
                if cur_upper not in {"ITEM", "DESCRIPTION", "QTY", "UNIT PRICE", "TOTAL"}:
                    desc_parts.append(cur)
                j += 1

            if qty:
                order += 1
                rows.append(
                    QuoteLine(
                        order=order,
                        item=item,
                        description=clean_text(" ".join(desc_parts)),
                        qty=safe_float(qty) if safe_float(qty) is not None else qty,
                        unit_price=unit,
                        total_price=total,
                        page=page_index,
                        notes=notes,
                    )
                )
                i = j
            else:
                i += 1

    # Detect FM1H replacement directly from full text if parser missed page note.
    for q in rows:
        if norm(q.item) == "FM1H" and not q.notes:
            if re.search(r"FM1H\s+Replaced with:\s+HF2,\s*DJTE-95L,\s*DJTE-95R", full_text, re.I | re.S):
                q.notes = "Replaced with: HF2, DJTE-95L, DJTE-95R"

    # Renumber after any cleanup.
    for idx, q in enumerate(rows, start=1):
        q.order = idx
    return meta, rows, full_text

def extract_quote_meta(text: str) -> Dict[str, Any]:
    invoice_no = ""
    invoice_date = ""
    customer = ""
    country = ""

    m_no = re.search(r"Invoice Number:\s*(\S+)", text, re.I)
    if m_no:
        invoice_no = m_no.group(1).strip()
    m_date = re.search(r"Invoice Date:\s*([0-9/\-]+)", text, re.I)
    if m_date:
        invoice_date = m_date.group(1).strip()

    m_bill = re.search(r"Bill to\s+(.+?)\s+\d", text, re.I | re.S)
    if m_bill:
        customer = clean_text(m_bill.group(1)).split(" ")[0:5]
        customer = " ".join(customer).replace("Ship To", "").strip()
    # More reliable for this quote: line after Bill to.
    lines = [clean_text(x) for x in text.splitlines() if clean_text(x)]
    for idx, line in enumerate(lines):
        if line.lower() == "bill to" and idx + 1 < len(lines):
            customer = lines[idx + 1]
            break
    for c in ["USA", "US", "CANADA", "CA"]:
        if re.search(rf"\b{re.escape(c)}\b", text, re.I):
            country = "USA" if c.upper() in {"US", "USA"} else c.upper()
            if c.upper() == "USA":
                break
    return {
        "customer": customer,
        "invoice_number": invoice_no,
        "invoice_date": invoice_date,
        "country": country,
    }


def quote_lines_to_df(lines: List[QuoteLine]) -> pd.DataFrame:
    return pd.DataFrame([asdict(x) for x in lines]).rename(columns={
        "order": "Order",
        "item": "Item",
        "description": "Description",
        "qty": "Qty",
        "unit_price": "Unit Price",
        "total_price": "Total Price",
        "page": "Page",
        "notes": "Notes",
    })


def generate_schedule(
    quote_lines: List[QuoteLine],
    master: MasterList,
    quickbooks: Optional[QuickBooksCatalog] = None,
    alias_df: Optional[pd.DataFrame] = None,
    replacement_df: Optional[pd.DataFrame] = None,
    special_df: Optional[pd.DataFrame] = None,
    ignore_df: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    alias_map = build_alias_map(alias_df if alias_df is not None else pd.DataFrame())
    replacement_map = build_replacement_map(replacement_df if replacement_df is not None else pd.DataFrame())
    special_map = build_special_map(special_df if special_df is not None else pd.DataFrame())
    ignore_map = build_ignore_map(ignore_df if ignore_df is not None else pd.DataFrame())

    schedule_rows: List[Dict[str, Any]] = []
    review_rows: List[Dict[str, Any]] = []
    group_seq = 1

    def next_group_id() -> str:
        nonlocal group_seq
        gid = f"G{group_seq:03d}"
        group_seq += 1
        return gid

    for q in quote_lines:
        q_norm = norm(q.item)

        if q_norm in ignore_map:
            review_rows.append(review_row(q, "Known Non-Schedule", ignore_map[q_norm], ""))
            continue

        special = special_map.get(q_norm)
        if special:
            review_rows.append(review_row(q, "Special Rule Requires Review", special.get("note") or special.get("action"), ""))
            continue

        # Replacements are never applied silently. They are suggestions for an
        # engineer to approve in the Review screen.
        replacement_entries = replacement_map.get(q_norm)
        if not replacement_entries and "REPLACED WITH" in q.notes.upper():
            after = q.notes.split(":", 1)[-1]
            replacement_entries = [(x.strip(), x.strip()) for x in after.split(",") if x.strip()]

        if replacement_entries:
            candidates = ", ".join(x[0] for x in replacement_entries)
            review_rows.append(review_row(q, "Replacement Requires Review", f"Suggested replacement(s): {candidates}", candidates))
            continue

        quantity = safe_float(q.qty)
        if quantity is not None and not is_whole_number(quantity):
            review_rows.append(review_row(
                q,
                "Non-Schedule Quantity / Review",
                "Quantity is not a whole equipment count. Treat as pricing/material quantity unless an engineer approves otherwise.",
                q.item,
            ))
            continue

        approved_alias = alias_map.get(q_norm)
        canonical_alias = CANONICAL_CODE_ALIASES.get(q_norm)
        lookup_item = approved_alias or canonical_alias or q.item
        if approved_alias:
            match_prefix = "Alias"
        elif canonical_alias:
            match_prefix = "Canonical Rename"
        else:
            match_prefix = "Exact"
        instance_count = qty_to_instances(q.qty)
        q_instance = QuoteLine(
            order=q.order,
            item=q.item,
            description=q.description,
            qty=1,
            unit_price=q.unit_price,
            total_price=q.total_price,
            page=q.page,
            notes=q.notes,
        )
        rows_before = len(schedule_rows)
        reviews_before = len(review_rows)
        add_lookup_to_schedule(
            schedule_rows,
            review_rows,
            master,
            quickbooks,
            q_instance,
            lookup_item,
            shown_quote_item=q.item,
            match_type_prefix=match_prefix,
            group_id=next_group_id(),
        )

        if len(schedule_rows) == rows_before:
            # Preserve the actual quote quantity on the single review record.
            for review in review_rows[reviews_before:]:
                review["Qty"] = q.qty
            continue

        generated_rows = schedule_rows[rows_before:]
        quickbooks_fallback_block = bool(generated_rows) and all(
            clean_text(row.get("Match Type")).startswith("QuickBooks ")
            for row in generated_rows
        )
        for row in generated_rows:
            row["Instance"] = 1
            row["Source Quote Qty"] = q.qty
            if quickbooks_fallback_block:
                # QuickBooks includes material/per-foot items. Keep their exact
                # quote quantity on one standalone row so a 500-foot item does
                # not create 500 browser and Excel rows.
                row["Qty"] = q.qty

        if quickbooks_fallback_block:
            continue

        # Whole equipment quantities repeat the complete parent/child group.
        for instance in range(2, instance_count + 1):
            block_start = len(schedule_rows)
            add_lookup_to_schedule(
                schedule_rows,
                review_rows,
                master,
                quickbooks,
                q_instance,
                lookup_item,
                shown_quote_item=q.item,
                match_type_prefix=match_prefix,
                group_id=next_group_id(),
            )
            for row in schedule_rows[block_start:]:
                row["Instance"] = instance
                row["Source Quote Qty"] = q.qty

    schedule_df = refresh_requirement_status(
        refresh_schedule_calculations(pd.DataFrame(schedule_rows, columns=SCHEDULE_COLUMNS)),
        master,
    )
    review_df = pd.DataFrame(review_rows, columns=REVIEW_COLUMNS)
    return schedule_df, review_df


def resolve_truncated_lookup(master: MasterList, q: QuoteLine, lookup_item: str) -> Tuple[Optional[str], str]:
    """Try to recover a truncated QuickBooks/PDF item code using Master List part numbers.

    This is still part-number based. It does not use description similarity.
    It only uses the visible part-number prefix before the ellipsis, and
    optionally verifies a full candidate part number if it appears in the
    quote description.
    """
    prefix = norm(lookup_item).split("...", 1)[0]
    if not prefix:
        return None, "No visible part-number prefix before ellipsis."

    candidates = master.candidate_part_numbers_starting_with(prefix)
    if not candidates:
        return None, "No Master List part number starts with the visible truncated prefix."

    desc_norm = norm(q.description)
    desc_hits = [c for c in candidates if norm(c) and norm(c) in desc_norm]
    if len(desc_hits) == 1:
        return desc_hits[0], "Recovered from full part number appearing inside quote description."

    if len(candidates) == 1:
        return candidates[0], "Recovered from unique Master List part number starting with truncated prefix."

    return None, "Multiple Master List part numbers start with truncated prefix: " + ", ".join(candidates[:10])


def add_lookup_to_schedule(
    schedule_rows: List[Dict[str, Any]],
    review_rows: List[Dict[str, Any]],
    master: MasterList,
    quickbooks: Optional[QuickBooksCatalog],
    q: QuoteLine,
    lookup_item: str,
    shown_quote_item: str,
    match_type_prefix: str,
    group_id: str,
) -> None:
    if "..." in lookup_item:
        resolved, reason = resolve_truncated_lookup(master, q, lookup_item)
        if resolved:
            review_start = len(review_rows)
            matched, terminal_review = try_add_exact_or_component_match(
                schedule_rows,
                review_rows,
                master,
                q,
                resolved,
                shown_quote_item,
                f"{match_type_prefix} Truncated Recovery",
                group_id,
            )
            if matched:
                return
            master_review = review_rows[review_start:]
            del review_rows[review_start:]
            qb_matched, qb_failure = try_add_quickbooks_fallback(
                schedule_rows, quickbooks, q, shown_quote_item, resolved, group_id,
            )
            if qb_matched:
                return
            if terminal_review:
                for review in master_review:
                    review["Details"] = (
                        clean_text(review.get("Details"))
                        + " QuickBooks fallback did not provide one unique exact item-number match."
                    ).strip()
                review_rows.extend(master_review)
                return
            if qb_failure == "ambiguous":
                review_rows.append(review_row(
                    q, "Ambiguous QuickBooks Match",
                    "QuickBooks has multiple different descriptions for this exact item number.",
                    resolved,
                ))
                return
            review_rows.append(review_row(q, "Truncated Item", f"{reason} Recovered code did not produce a safe Master match.", resolved))
            return
        qb_matched, qb_failure = try_add_quickbooks_fallback(
            schedule_rows, quickbooks, q, shown_quote_item, shown_quote_item, group_id,
        )
        if qb_matched:
            return
        if qb_failure == "ambiguous":
            review_rows.append(review_row(
                q, "Ambiguous QuickBooks Match",
                "QuickBooks has multiple different descriptions for this exact item number.",
                shown_quote_item,
            ))
            return
        review_rows.append(review_row(q, "Truncated Item", reason, lookup_item))
        return

    review_start = len(review_rows)
    matched, terminal_review = try_add_exact_or_component_match(
        schedule_rows, review_rows, master, q, lookup_item, shown_quote_item, match_type_prefix, group_id
    )
    if matched:
        return

    master_review = review_rows[review_start:]
    del review_rows[review_start:]
    qb_matched, qb_failure = try_add_quickbooks_fallback(
        schedule_rows, quickbooks, q, shown_quote_item, lookup_item, group_id,
    )
    if qb_matched:
        return
    if terminal_review:
        for review in master_review:
            review["Details"] = (
                clean_text(review.get("Details"))
                + " QuickBooks fallback did not provide one unique exact item-number match."
            ).strip()
        review_rows.extend(master_review)
        return
    if qb_failure == "ambiguous":
        review_rows.append(review_row(
            q,
            "Ambiguous QuickBooks Match",
            "QuickBooks has multiple different descriptions for this exact item number.",
            lookup_item,
        ))
        return

    review_rows.append(review_row(
        q,
        "Unmatched",
        "No safe Master List match or unique exact QuickBooks item-number match exists. No description guess was used.",
        lookup_item,
    ))


def try_add_quickbooks_fallback(
    schedule_rows: List[Dict[str, Any]],
    quickbooks: Optional[QuickBooksCatalog],
    q: QuoteLine,
    shown_quote_item: str,
    lookup_item: str,
    group_id: str,
) -> Tuple[bool, str]:
    """Use QuickBooks only after the Master List cannot produce a safe row."""
    if quickbooks is None:
        return False, "none"

    candidates = [shown_quote_item]
    if norm(lookup_item) != norm(shown_quote_item):
        candidates.append(lookup_item)

    ambiguous_matches: List[QuickBooksItem] = []
    for candidate in candidates:
        status, matches = quickbooks.find_exact(candidate)
        if status in {"single", "identical_duplicate"} and matches:
            match_type = "QuickBooks Exact Fallback"
            if status == "identical_duplicate":
                match_type += " (Identical Duplicate Auto-Resolved)"
            schedule_rows.append(matches[0].to_schedule_row(
                order=str(q.order),
                group_id=group_id,
                qty=q.qty,
                quote_item=shown_quote_item,
                match_type=match_type,
            ))
            return True, ""
        if status == "ambiguous":
            ambiguous_matches.extend(matches)

    return False, "ambiguous" if ambiguous_matches else "none"


def try_add_exact_or_component_match(
    schedule_rows: List[Dict[str, Any]],
    review_rows: List[Dict[str, Any]],
    master: MasterList,
    q: QuoteLine,
    lookup_item: str,
    shown_quote_item: str,
    match_type_prefix: str,
    group_id: str,
) -> Tuple[bool, bool]:
    """Try parent/component match. Returns (matched, terminal_review).

    terminal_review=True means a match was found but it is ambiguous and should not fall back.
    """
    # WA1-SK-2018 exists in several historical assemblies. Their component
    # requirements are identical, but their parent packages differ, so a normal
    # parent lookup is intentionally ambiguous. For a direct quote call, build
    # the approved minimal kit + required panel pair deterministically.
    if norm(lookup_item) == WRAP_AIR_ASSIST_KIT:
        kit_status, kit_matches = master.find_component(WRAP_AIR_ASSIST_KIT)
        if kit_status in {"single", "identical_duplicate"} and kit_matches:
            kit_source = kit_matches[0]
            root = kit_source.to_schedule_row(
                order=str(q.order),
                group_id=group_id,
                qty=q.qty,
                quote_item=shown_quote_item,
                quote_description=q.description,
                match_type=f"{match_type_prefix} Approved Kit",
            )
            root["Item"] = WRAP_AIR_ASSIST_KIT
            root["Nested Code"] = "Parent"
            root["Depth"] = 0
            root["Is Assembly Root"] = True
            matched_rows = ensure_wrap_air_assist_panel(
                [root], master, WRAP_AIR_ASSIST_KIT, str(q.order), group_id,
                q.qty, shown_quote_item, match_type_prefix,
            )
            if any(norm(row.get("Item")) == WRAP_AIR_ASSIST_PANEL for row in matched_rows):
                schedule_rows.extend(matched_rows)
                return True, False

    parent_status, parent_matches = master.find_parent(lookup_item)
    if parent_status in {"single", "identical_duplicate"}:
        parent = parent_matches[0]
        group_rows = master.group_for_parent(parent)
        matched_rows: List[Dict[str, Any]] = []
        for row in group_rows:
            mt = f"{match_type_prefix} Parent" if row.is_parent else f"{match_type_prefix} Child"
            if parent_status == "identical_duplicate" and row.is_parent:
                mt += " (Identical Duplicate Auto-Resolved)"
            matched_rows.append(
                row.to_schedule_row(
                    order=str(q.order),
                    group_id=group_id,
                    qty=q.qty,
                    quote_item=shown_quote_item,
                    quote_description=q.description,
                    match_type=mt,
                )
            )
        matched_rows = ensure_wrap_air_assist_panel(
            matched_rows, master, parent.part_number, str(q.order), group_id,
            q.qty, shown_quote_item, match_type_prefix,
        )
        schedule_rows.extend(matched_rows)
        return True, False
    if parent_status == "ambiguous":
        candidates = ", ".join([f"row {m.excel_row}" for m in parent_matches[:10]])
        review_rows.append(review_row(q, "Ambiguous Parent Match", f"Multiple parent groups found: {candidates}", lookup_item))
        return False, True

    comp_status, comp_matches = master.find_component(lookup_item)
    if comp_status in {"single", "identical_duplicate"}:
        comp = comp_matches[0]
        mt = f"{match_type_prefix} Component"
        if comp_status == "identical_duplicate":
            mt += " (Identical Duplicate Auto-Resolved)"
        schedule_rows.append(
            comp.to_schedule_row(
                order=str(q.order),
                group_id=group_id,
                qty=q.qty,
                quote_item=shown_quote_item,
                quote_description=q.description,
                match_type=mt,
            )
        )
        return True, False
    if comp_status == "ambiguous":
        candidates = ", ".join([f"row {m.excel_row}" for m in comp_matches[:10]])
        review_rows.append(review_row(q, "Ambiguous Component Match", f"Multiple component rows found: {candidates}", lookup_item))
        return False, True

    return False, False


def review_row(q: QuoteLine, issue_type: str, details: str, candidate_lookup: str) -> Dict[str, Any]:
    return {
        "Order": q.order,
        "Item": q.item,
        "Description": q.description,
        "Qty": q.qty,
        "Issue Type": issue_type,
        "Details": details,
        "Candidate Lookup": candidate_lookup,
    }


def build_parent_catalog_df(master: MasterList) -> pd.DataFrame:
    """Return parent/bold rows from the Master List for the web edit screen."""
    records: List[Dict[str, Any]] = []
    seen: Dict[str, int] = {}
    for row in master.rows:
        if not row.is_parent or not row.part_number:
            continue
        item_key = norm(row.part_number)
        seen[item_key] = seen.get(item_key, 0) + 1
        records.append({
            "Excel Row": row.excel_row,
            "Item": row.part_number,
            "Description": row.description,
            "Master Group ID": row.group_id,
            "Duplicate Count For Item": seen[item_key],
        })
    df = pd.DataFrame(records)
    if not df.empty:
        df["Search Label"] = df.apply(
            lambda r: f"{r['Item']} | {str(r['Description'])[:90]} | Excel row {r['Excel Row']}",
            axis=1,
        )
    return df


def build_component_catalog_df(master: MasterList) -> pd.DataFrame:
    """Return every exact Master row for component-level insert/replace actions."""
    records: List[Dict[str, Any]] = []
    for row in master.rows:
        if not row.part_number:
            continue
        records.append({
            "Excel Row": row.excel_row,
            "Item": row.part_number,
            "Description": row.description,
            "Master Group ID": row.group_id,
            "Is Master Parent": bool(row.is_parent),
            "Is Master Assembly": bool(row.is_parent or row.is_bold),
            "Requirements": " | ".join(
                f"{label}: {clean_text(getattr(row, key))}"
                for label, key in [
                    ("HP", "hp"), ("PHASE", "phase"), ("VOLTS", "volts"),
                    ("AMPS", "amps"), ("CB", "cb"), ("AIR PORT", "air_port"),
                    ("COLD WATER", "cold_water"), ("HOT WATER", "hot_water"),
                    ("RECLAIM WATER", "reclaim_water"), ("GAS BTUH", "gas_btuh"),
                ]
                if clean_text(getattr(row, key)) not in {"", "-"}
            ) or "No electrical or utility requirements listed",
        })
    df = pd.DataFrame(records)
    if not df.empty:
        df["Search Label"] = df.apply(
            lambda r: (
                f"{r['Item']} | {str(r['Description'])[:72]} | "
                f"Master row {r['Excel Row']}"
                + (" | stored assembly" if bool(r["Is Master Assembly"]) else " | component")
            ),
            axis=1,
        )
    return df


def build_quickbooks_catalog_df(quickbooks: Optional[QuickBooksCatalog]) -> pd.DataFrame:
    """Return a compact searchable item/description table for the web app."""
    if quickbooks is None:
        return pd.DataFrame(columns=[
            "QuickBooks Row", "Item", "Description", "Primary Category",
            "Schedule Status", "Active Status", "QuickBooks Category", "Requirements",
        ])
    records = [{
        "QuickBooks Row": item.excel_row,
        "Item": item.item_number,
        "Description": item.description,
        "Primary Category": item.primary_category,
        "Schedule Status": item.schedule_status,
        "Active Status": item.active_status,
        "QuickBooks Category": item.quickbooks_category,
        "Requirements": "No engineering requirements in QuickBooks",
    } for item in quickbooks.items]
    return pd.DataFrame(records)


def ensure_wrap_air_assist_panel(
    schedule_rows: List[Dict[str, Any]],
    master: MasterList,
    root_item: Any,
    order: str,
    group_id: str,
    qty: Any,
    quote_item: str,
    match_type_prefix: str,
) -> List[Dict[str, Any]]:
    """Ensure the approved panel is a direct child of the wrap-air kit.

    Older Master groups may already contain WA1-SKP, while the newer standalone
    WA1-SK-2018 group contains only solenoids. This rule adds the canonical
    panel exactly once and nests the existing children beneath it.
    """
    rows = [dict(row) for row in schedule_rows]
    if norm(root_item) != WRAP_AIR_ASSIST_KIT:
        return rows
    if any(norm(row.get("Item")) == WRAP_AIR_ASSIST_PANEL for row in rows):
        return rows

    panel_status, panel_matches = master.find_component(WRAP_AIR_ASSIST_PANEL)
    if panel_status not in {"single", "identical_duplicate"}:
        # Compatibility with a pre-rename Master workbook.
        panel_status, panel_matches = master.find_component("WA1-SKP")
    if panel_status not in {"single", "identical_duplicate"} or not panel_matches:
        return rows

    panel_source = panel_matches[0]
    panel_row = panel_source.to_schedule_row(
        order=order,
        group_id=group_id,
        qty=qty,
        quote_item=quote_item,
        quote_description="",
        match_type=f"{match_type_prefix} Required Child",
    )
    panel_row["Item"] = WRAP_AIR_ASSIST_PANEL
    panel_row["Nested Code"] = "A"
    panel_row["Depth"] = 1
    panel_row["Is Assembly Root"] = True

    if not rows:
        return [panel_row]
    root = rows[0]
    descendants: List[Dict[str, Any]] = []
    for row in rows[1:]:
        nested = clean_text(row.get("Nested Code")).upper()
        if nested and nested.lower() != "parent":
            row["Nested Code"] = "A" + nested
            row["Depth"] = max(2, int(row.get("Depth", 1) or 1) + 1)
        descendants.append(row)
    return [root, panel_row, *descendants]


def schedule_rows_for_parent_excel_row(
    master: MasterList,
    excel_row: int,
    qty: Any = 1,
    group_id: str = "G000",
    order: str = "",
) -> List[Dict[str, Any]]:
    """Create schedule rows for one exact parent row selected from the Master Library."""
    parent = next((r for r in master.rows if r.excel_row == int(excel_row) and r.is_parent), None)
    if parent is None:
        raise ValueError(f"No parent row found at Master List Excel row {excel_row}.")
    out: List[Dict[str, Any]] = []
    for row in master.group_for_parent(parent):
        mt = "Manual Add Parent" if row.is_parent else "Manual Add Child"
        out.append(
            row.to_schedule_row(
                order=order,
                group_id=group_id,
                qty=qty,
                quote_item=parent.part_number,
                quote_description=parent.description,
                match_type=mt,
            )
        )
    return ensure_wrap_air_assist_panel(
        out, master, parent.part_number, clean_text(order), group_id, qty,
        parent.part_number, "Manual Add",
    )


def schedule_rows_for_master_subtree_excel_row(
    master: MasterList,
    excel_row: int,
    qty: Any = 1,
    group_id: str = "G000",
    order: str = "",
) -> List[Dict[str, Any]]:
    """Create the stored subtree rooted at any bold Master row.

    A top-level Master parent returns its whole group. A bold nested assembly
    returns itself plus only the following descendants under its code prefix.
    """
    selected = next((r for r in master.rows if r.excel_row == int(excel_row)), None)
    if selected is None or not (selected.is_parent or selected.is_bold):
        raise ValueError(f"Master row {excel_row} is not a stored bold assembly.")
    group_rows = master.groups.get(selected.group_id, [])
    if selected.is_parent:
        subtree = master.group_for_parent(selected)
        selected_prefix = ""
    else:
        selected_prefix = clean_text(selected.nested_code).upper()
        start = group_rows.index(selected)
        subtree = [selected]
        for candidate in group_rows[start + 1:]:
            code = clean_text(candidate.nested_code).upper()
            if code.startswith(selected_prefix) and len(code) > len(selected_prefix):
                subtree.append(candidate)
            else:
                break

    out: List[Dict[str, Any]] = []
    for position, row in enumerate(subtree):
        schedule_row = row.to_schedule_row(
            order=order,
            group_id=group_id,
            qty=qty,
            quote_item=selected.part_number,
            quote_description=selected.description,
            match_type="Manual Add Stored Assembly" if position == 0 else "Manual Add Stored Child",
        )
        if position == 0:
            schedule_row["Nested Code"] = "Parent"
            schedule_row["Is Assembly Root"] = True
        elif selected_prefix:
            schedule_row["Nested Code"] = clean_text(row.nested_code)[len(selected_prefix):]
        out.append(schedule_row)
    return ensure_wrap_air_assist_panel(
        out, master, selected.part_number, clean_text(order), group_id, qty,
        selected.part_number, "Manual Add",
    )


def schedule_row_for_master_excel_row(
    master: MasterList,
    excel_row: int,
    group_id: str = "G000",
    order: str = "",
) -> Dict[str, Any]:
    """Create one project row from one exact Master row, inheriting requirements."""
    selected = next((r for r in master.rows if r.excel_row == int(excel_row)), None)
    if selected is None:
        raise ValueError(f"No Master List row found at Excel row {excel_row}.")
    row = selected.to_schedule_row(
        order=order,
        group_id=group_id,
        qty=1,
        quote_item=selected.part_number,
        quote_description=selected.description,
        match_type="Manual Add From Master",
    )
    # A component inserted on its own is not automatically a top-level assembly.
    row["Nested Code"] = "Parent"
    row["Parent Row ID"] = ""
    row["Depth"] = 0
    return row


def schedule_row_for_quickbooks_excel_row(
    quickbooks: QuickBooksCatalog,
    excel_row: int,
    group_id: str = "G000",
    order: str = "",
) -> Dict[str, Any]:
    """Create one standalone project row from a selected QuickBooks item."""
    selected = quickbooks.row_index.get(int(excel_row))
    if selected is None:
        raise ValueError(f"No QuickBooks item found at Unique Items row {excel_row}.")
    return selected.to_schedule_row(
        order=order,
        group_id=group_id,
        qty=1,
        quote_item=selected.item_number,
        match_type="Manual Add From QuickBooks",
    )


def _excel_alpha(position: int) -> str:
    """Return 1=A, 26=Z, 27=AA for sibling labels."""
    position = max(1, int(position))
    letters = ""
    while position:
        position, remainder = divmod(position - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _next_row_id(schedule_df: pd.DataFrame) -> str:
    maximum = 0
    if schedule_df is not None and "Row ID" in schedule_df.columns:
        for value in schedule_df["Row ID"].dropna().astype(str):
            match = re.fullmatch(r"R(\d+)", value.strip(), flags=re.IGNORECASE)
            if match:
                maximum = max(maximum, int(match.group(1)))
    return f"R{maximum + 1:06d}"


def ensure_schedule_hierarchy(schedule_df: pd.DataFrame) -> pd.DataFrame:
    """Create stable row IDs, infer legacy parents, and regenerate outline codes.

    Row ID is permanent inside a draft. Order, Group ID, Parent Group ID,
    Depth, and Nested Code are display fields derived from Parent Row ID and
    current sibling order. This prevents stale A/AA/AAA labels after editing.
    """
    if schedule_df is None:
        return pd.DataFrame(columns=SCHEDULE_COLUMNS)
    df = schedule_df.copy().reset_index(drop=True)
    if df.empty:
        for col in SCHEDULE_COLUMNS:
            if col not in df.columns:
                df[col] = pd.Series(dtype="object")
        return df[SCHEDULE_COLUMNS]

    defaults: Dict[str, Any] = {
        "Row ID": "", "Parent Row ID": "", "Depth": 0,
        "Requirements Source": "", "Requirements Status": "Custom / project-only",
        "Bold": False, "Italic": False, "Underline": False, "Highlight": False,
        "Review Added": False,
        "Parent Group ID": "",
    }
    for col in SCHEDULE_COLUMNS:
        if col not in df.columns:
            df[col] = defaults.get(col, "")

    # Existing stable IDs mean blank Parent Row IDs are intentional top-level
    # roots. Only legacy frames that arrived without IDs need code/group-based
    # parent inference.
    had_stable_row_ids = df["Row ID"].astype(str).str.strip().any()

    # Allocate or repair stable IDs without using the dataframe index as identity.
    used: set[str] = set()
    next_number = 1
    row_ids: List[str] = []
    for value in df["Row ID"].tolist():
        candidate = clean_text(value)
        if candidate and candidate not in used:
            row_ids.append(candidate)
            used.add(candidate)
            continue
        while f"R{next_number:06d}" in used:
            next_number += 1
        candidate = f"R{next_number:06d}"
        next_number += 1
        row_ids.append(candidate)
        used.add(candidate)
    df["Row ID"] = row_ids

    # Infer Parent Row ID once for legacy drafts using their existing group/code.
    valid_ids = set(row_ids)
    needs_inference = (
        not had_stable_row_ids
        and not df["Parent Row ID"].astype(str).str.strip().any()
    )
    if needs_inference:
        inferred: Dict[int, str] = {}
        for _, block in df.groupby(df["Group ID"].astype(str), sort=False):
            root_idx = block.index[0]
            for idx in block.index:
                if clean_text(df.at[idx, "Nested Code"]).lower() == "parent":
                    root_idx = idx
                    break
            root_id = df.at[root_idx, "Row ID"]
            code_to_id: Dict[str, str] = {}
            inferred[root_idx] = ""
            for idx in block.index:
                if idx == root_idx:
                    continue
                code = clean_text(df.at[idx, "Nested Code"]).upper()
                parent_id = root_id
                if code:
                    # Existing AVW codes use the path prefix: AA belongs to A.
                    prefix = code[:-1]
                    if prefix and prefix in code_to_id:
                        parent_id = code_to_id[prefix]
                    code_to_id[code] = df.at[idx, "Row ID"]
                inferred[idx] = parent_id
        df["Parent Row ID"] = [inferred.get(idx, "") for idx in df.index]

    # Invalid/self parent links become top-level rows. Cycles are also broken.
    parent_by_id: Dict[str, str] = {}
    for _, row in df.iterrows():
        rid = clean_text(row["Row ID"])
        parent = clean_text(row["Parent Row ID"])
        if parent not in valid_ids or parent == rid:
            parent = ""
        cursor = parent
        visited = {rid}
        cycle = False
        while cursor:
            if cursor in visited:
                cycle = True
                break
            visited.add(cursor)
            prior = parent_by_id.get(cursor, "")
            cursor = prior
        parent_by_id[rid] = "" if cycle else parent
    df["Parent Row ID"] = df["Row ID"].map(parent_by_id).fillna("")

    source_order = {rid: pos for pos, rid in enumerate(df["Row ID"].tolist())}
    children: Dict[str, List[str]] = {}
    for rid in df["Row ID"].tolist():
        children.setdefault(parent_by_id.get(rid, ""), []).append(rid)
    for child_ids in children.values():
        child_ids.sort(key=lambda rid: source_order[rid])

    records = {str(row["Row ID"]): row.to_dict() for _, row in df.iterrows()}
    ordered_rows: List[Dict[str, Any]] = []

    def visit(rid: str, root_number: int, group_id: str, code: str, depth: int) -> None:
        record = records[rid]
        record["Order"] = str(root_number)
        record["Group ID"] = group_id
        record["Parent Group ID"] = "" if depth == 0 else group_id
        record["Depth"] = depth
        record["Nested Code"] = "Parent" if depth == 0 else code
        if depth == 0:
            record["Is Assembly Root"] = bool(record.get("Is Assembly Root"))
        ordered_rows.append(record)
        for sibling_number, child_id in enumerate(children.get(rid, []), start=1):
            child_code = code + _excel_alpha(sibling_number)
            visit(child_id, root_number, group_id, child_code, depth + 1)

    roots = children.get("", [])
    for root_number, root_id in enumerate(roots, start=1):
        visit(root_id, root_number, f"G{root_number:03d}", "", 0)

    # Any record omitted because of malformed legacy links is safely promoted.
    emitted = {row["Row ID"] for row in ordered_rows}
    for root_id in [rid for rid in df["Row ID"].tolist() if rid not in emitted]:
        root_number = len([r for r in ordered_rows if r["Nested Code"] == "Parent"]) + 1
        parent_by_id[root_id] = ""
        visit(root_id, root_number, f"G{root_number:03d}", "", 0)

    out = pd.DataFrame(ordered_rows)
    for col in SCHEDULE_COLUMNS:
        if col not in out.columns:
            out[col] = defaults.get(col, "")
    return out[SCHEDULE_COLUMNS]


def next_group_id_from_df(schedule_df: pd.DataFrame) -> str:
    if schedule_df is None or schedule_df.empty or "Group ID" not in schedule_df.columns:
        return "G001"
    max_num = 0
    for gid in schedule_df["Group ID"].dropna().astype(str):
        m = re.search(r"(\d+)$", gid)
        if m:
            max_num = max(max_num, int(m.group(1)))
    return f"G{max_num + 1:03d}"


def group_sequence(schedule_df: pd.DataFrame) -> List[str]:
    if schedule_df is None or schedule_df.empty or "Group ID" not in schedule_df.columns:
        return []
    seq: List[str] = []
    for gid in schedule_df["Group ID"].astype(str):
        if gid and gid not in seq:
            seq.append(gid)
    return seq


def renumber_schedule_orders(schedule_df: pd.DataFrame) -> pd.DataFrame:
    """Compatibility wrapper for the hierarchy-based renumbering engine."""
    return ensure_schedule_hierarchy(schedule_df)


def insert_rows_relative(
    schedule_df: pd.DataFrame,
    new_rows_df: pd.DataFrame,
    mode: str = "bottom",
    target_group_id: str = "",
) -> pd.DataFrame:
    """Legacy group insertion wrapper backed by row-tree insertion."""
    df = ensure_schedule_hierarchy(schedule_df)
    target_row_id = ""
    if target_group_id and not df.empty:
        match = df[df["Group ID"].astype(str) == str(target_group_id)]
        if not match.empty:
            target_row_id = str(match.iloc[0]["Row ID"])
    return insert_schedule_rows(df, new_rows_df, mode=mode, target_row_id=target_row_id)


def _subtree_ids(schedule_df: pd.DataFrame, row_id: str) -> List[str]:
    parent_map = schedule_df.set_index("Row ID")["Parent Row ID"].astype(str).to_dict()
    descendants = {str(row_id)}
    changed = True
    while changed:
        changed = False
        for candidate, parent in parent_map.items():
            if candidate not in descendants and parent in descendants:
                descendants.add(candidate)
                changed = True
    return [rid for rid in schedule_df["Row ID"].astype(str).tolist() if rid in descendants]


def _selection_root_ids(schedule_df: pd.DataFrame, row_ids: Iterable[str]) -> List[str]:
    """Return selected rows in schedule order, excluding selected descendants.

    When both a parent and one of its children are selected, the parent owns
    the operation so the child subtree is not moved or deleted twice.
    """
    df = ensure_schedule_hierarchy(schedule_df)
    selected = {
        clean_text(row_id)
        for row_id in row_ids
        if clean_text(row_id)
    }
    valid_ids = set(df["Row ID"].astype(str))
    selected &= valid_ids
    if not selected:
        return []
    parent_map = df.set_index("Row ID")["Parent Row ID"].astype(str).to_dict()
    roots: List[str] = []
    for row_id in df["Row ID"].astype(str).tolist():
        if row_id not in selected:
            continue
        ancestor = clean_text(parent_map.get(row_id, ""))
        owned_by_selected_ancestor = False
        while ancestor:
            if ancestor in selected:
                owned_by_selected_ancestor = True
                break
            ancestor = clean_text(parent_map.get(ancestor, ""))
        if not owned_by_selected_ancestor:
            roots.append(row_id)
    return roots


def selected_subtree_ids(schedule_df: pd.DataFrame, row_ids: Iterable[str]) -> List[str]:
    """Return the exact schedule-order rows affected by subtree operations."""
    df = ensure_schedule_hierarchy(schedule_df)
    affected: set[str] = set()
    for root_id in _selection_root_ids(df, row_ids):
        affected.update(_subtree_ids(df, root_id))
    return [row_id for row_id in df["Row ID"].astype(str).tolist() if row_id in affected]


def subtree_descendant_counts(schedule_df: pd.DataFrame) -> Dict[str, int]:
    """Count every row's descendants in one hierarchy pass.

    This lets the web grid label parent groups without rebuilding a complete
    subtree separately for every displayed row during checkbox selection.
    """
    df = ensure_schedule_hierarchy(schedule_df)
    if df.empty:
        return {}
    row_ids = df["Row ID"].astype(str).tolist()
    parent_map = df.set_index("Row ID")["Parent Row ID"].astype(str).to_dict()
    counts = {row_id: 0 for row_id in row_ids}
    for row_id in row_ids:
        ancestor = clean_text(parent_map.get(row_id, ""))
        visited = {row_id}
        while ancestor and ancestor not in visited:
            visited.add(ancestor)
            if ancestor in counts:
                counts[ancestor] += 1
            ancestor = clean_text(parent_map.get(ancestor, ""))
    return counts


def _remap_new_row_ids(existing_df: pd.DataFrame, new_rows_df: pd.DataFrame) -> pd.DataFrame:
    new_df = ensure_schedule_hierarchy(new_rows_df)
    if new_df.empty:
        return new_df
    used = set(existing_df.get("Row ID", pd.Series(dtype=str)).astype(str).tolist())
    maximum = 0
    for rid in used:
        match = re.fullmatch(r"R(\d+)", rid, flags=re.IGNORECASE)
        if match:
            maximum = max(maximum, int(match.group(1)))
    mapping: Dict[str, str] = {}
    for old_id in new_df["Row ID"].astype(str).tolist():
        maximum += 1
        while f"R{maximum:06d}" in used:
            maximum += 1
        mapping[old_id] = f"R{maximum:06d}"
        used.add(mapping[old_id])
    new_df["Row ID"] = new_df["Row ID"].astype(str).map(mapping)
    new_df["Parent Row ID"] = new_df["Parent Row ID"].astype(str).map(mapping).fillna("")
    return new_df


def insert_schedule_rows(
    schedule_df: pd.DataFrame,
    new_rows_df: pd.DataFrame,
    mode: str = "bottom",
    target_row_id: str = "",
) -> pd.DataFrame:
    """Insert one row or a complete subtree at top/bottom or around a target."""
    df = ensure_schedule_hierarchy(schedule_df)
    if new_rows_df is None or new_rows_df.empty:
        return refresh_schedule_calculations(df)
    new_df = _remap_new_row_ids(df, new_rows_df)
    mode = clean_text(mode).lower() or "bottom"
    target_row_id = clean_text(target_row_id)
    valid_target = target_row_id in set(df.get("Row ID", pd.Series(dtype=str)).astype(str))

    roots = new_df[new_df["Parent Row ID"].astype(str).eq("")]["Row ID"].astype(str).tolist()
    insert_at = 0 if mode == "top" else len(df)
    if valid_target and mode in {"before", "after", "child"}:
        target_idx = int(df.index[df["Row ID"].astype(str).eq(target_row_id)][0])
        target_parent = clean_text(df.at[target_idx, "Parent Row ID"])
        if mode == "before":
            insert_at = target_idx
            for root in roots:
                new_df.loc[new_df["Row ID"].astype(str).eq(root), "Parent Row ID"] = target_parent
        elif mode == "after":
            subtree = _subtree_ids(df, target_row_id)
            insert_at = max(int(df.index[df["Row ID"].astype(str).eq(rid)][0]) for rid in subtree) + 1
            for root in roots:
                new_df.loc[new_df["Row ID"].astype(str).eq(root), "Parent Row ID"] = target_parent
        else:
            subtree = _subtree_ids(df, target_row_id)
            insert_at = max(int(df.index[df["Row ID"].astype(str).eq(rid)][0]) for rid in subtree) + 1
            for root in roots:
                new_df.loc[new_df["Row ID"].astype(str).eq(root), "Parent Row ID"] = target_row_id

    combined = pd.concat([df.iloc[:insert_at], new_df, df.iloc[insert_at:]], ignore_index=True)
    return refresh_schedule_calculations(ensure_schedule_hierarchy(combined))


def delete_group(schedule_df: pd.DataFrame, group_id: str) -> pd.DataFrame:
    if schedule_df is None or schedule_df.empty or not group_id:
        return schedule_df
    df = ensure_schedule_hierarchy(schedule_df)
    match = df[df["Group ID"].astype(str) == str(group_id)]
    if match.empty:
        return df
    return delete_row_subtree(df, str(match.iloc[0]["Row ID"]))


def delete_row_subtree(schedule_df: pd.DataFrame, row_id: str) -> pd.DataFrame:
    df = ensure_schedule_hierarchy(schedule_df)
    if row_id not in set(df["Row ID"].astype(str)):
        return df
    remove_ids = set(_subtree_ids(df, row_id))
    return refresh_schedule_calculations(
        ensure_schedule_hierarchy(df[~df["Row ID"].astype(str).isin(remove_ids)].reset_index(drop=True))
    )


def delete_row_subtrees(schedule_df: pd.DataFrame, row_ids: Iterable[str]) -> pd.DataFrame:
    """Delete several selected subtrees once, safely de-duplicating descendants."""
    df = ensure_schedule_hierarchy(schedule_df)
    remove_ids = set(selected_subtree_ids(df, row_ids))
    if not remove_ids:
        return df
    return refresh_schedule_calculations(
        ensure_schedule_hierarchy(df[~df["Row ID"].astype(str).isin(remove_ids)].reset_index(drop=True))
    )


def move_group(schedule_df: pd.DataFrame, group_id: str, direction: int) -> pd.DataFrame:
    """Move one parent block up/down while keeping its nested rows together."""
    df = ensure_schedule_hierarchy(schedule_df)
    match = df[df["Group ID"].astype(str) == str(group_id)]
    if match.empty:
        return df
    return move_row_subtree(df, str(match.iloc[0]["Row ID"]), direction)


def move_row_subtree(schedule_df: pd.DataFrame, row_id: str, direction: int) -> pd.DataFrame:
    """Move a row and all descendants one place among its siblings."""
    df = ensure_schedule_hierarchy(schedule_df)
    match = df[df["Row ID"].astype(str).eq(str(row_id))]
    if match.empty:
        return df
    parent_id = clean_text(match.iloc[0]["Parent Row ID"])
    siblings = df[df["Parent Row ID"].astype(str).eq(parent_id)]["Row ID"].astype(str).tolist()
    position = siblings.index(str(row_id))
    new_position = position + (-1 if int(direction) < 0 else 1)
    if new_position < 0 or new_position >= len(siblings):
        return df
    neighbor_id = siblings[new_position]
    selected_ids = _subtree_ids(df, str(row_id))
    neighbor_ids = _subtree_ids(df, neighbor_id)
    ids = df["Row ID"].astype(str).tolist()
    selected_start = ids.index(selected_ids[0])
    neighbor_start = ids.index(neighbor_ids[0])
    remaining = [rid for rid in ids if rid not in set(selected_ids + neighbor_ids)]
    insert_at = min(selected_start, neighbor_start)
    left = remaining[:insert_at]
    right = remaining[insert_at:]
    pair = selected_ids + neighbor_ids if int(direction) > 0 else selected_ids + neighbor_ids
    if int(direction) < 0:
        pair = selected_ids + neighbor_ids
    else:
        pair = neighbor_ids + selected_ids
    # For upward movement the selected block is placed before its previous sibling.
    if int(direction) < 0:
        pair = selected_ids + neighbor_ids
    ordered_ids = left + pair + right
    out = df.set_index("Row ID").loc[ordered_ids].reset_index()
    return refresh_schedule_calculations(ensure_schedule_hierarchy(out))


def move_row_subtrees(
    schedule_df: pd.DataFrame,
    row_ids: Iterable[str],
    target_row_id: str = "",
    mode: str = "after",
) -> pd.DataFrame:
    """Move selected subtree roots as one ordered block to a precise location.

    Supported modes are before, after, child, and bottom. Parent/child links of
    descendants remain intact; only the selected roots receive the destination
    parent. A target inside the moving selection is rejected safely.
    """
    df = ensure_schedule_hierarchy(schedule_df)
    source_roots = _selection_root_ids(df, row_ids)
    if not source_roots:
        return df
    moving_ids = selected_subtree_ids(df, source_roots)
    moving_set = set(moving_ids)
    target_row_id = clean_text(target_row_id)
    mode = clean_text(mode).lower() or "after"
    if mode not in {"before", "after", "child", "bottom"}:
        return df
    if target_row_id in moving_set:
        return df

    block = df[df["Row ID"].astype(str).isin(moving_set)].copy()
    remainder = df[~df["Row ID"].astype(str).isin(moving_set)].copy().reset_index(drop=True)
    remainder_ids = set(remainder["Row ID"].astype(str))

    insert_at = len(remainder)
    destination_parent = ""
    if mode != "bottom":
        if target_row_id not in remainder_ids:
            return df
        target_idx = int(remainder.index[remainder["Row ID"].astype(str).eq(target_row_id)][0])
        if mode == "before":
            insert_at = target_idx
            destination_parent = clean_text(remainder.at[target_idx, "Parent Row ID"])
        elif mode == "after":
            target_subtree = _subtree_ids(remainder, target_row_id)
            insert_at = max(
                int(remainder.index[remainder["Row ID"].astype(str).eq(row_id)][0])
                for row_id in target_subtree
            ) + 1
            destination_parent = clean_text(remainder.at[target_idx, "Parent Row ID"])
        else:
            target_subtree = _subtree_ids(remainder, target_row_id)
            insert_at = max(
                int(remainder.index[remainder["Row ID"].astype(str).eq(row_id)][0])
                for row_id in target_subtree
            ) + 1
            destination_parent = target_row_id

    root_mask = block["Row ID"].astype(str).isin(source_roots)
    block.loc[root_mask, "Parent Row ID"] = destination_parent
    combined = pd.concat(
        [remainder.iloc[:insert_at], block, remainder.iloc[insert_at:]],
        ignore_index=True,
    )
    return refresh_schedule_calculations(ensure_schedule_hierarchy(combined))


def reparent_row(schedule_df: pd.DataFrame, row_id: str, new_parent_row_id: str) -> pd.DataFrame:
    """Move a complete subtree to the end of a new parent's children."""
    df = ensure_schedule_hierarchy(schedule_df)
    row_id = clean_text(row_id)
    new_parent_row_id = clean_text(new_parent_row_id)
    ids = set(df["Row ID"].astype(str))
    if row_id not in ids or new_parent_row_id not in ids or row_id == new_parent_row_id:
        return df
    subtree_ids = _subtree_ids(df, row_id)
    if new_parent_row_id in subtree_ids:
        return df
    block = df[df["Row ID"].astype(str).isin(subtree_ids)].copy()
    remainder = df[~df["Row ID"].astype(str).isin(subtree_ids)].copy().reset_index(drop=True)
    block.loc[block["Row ID"].astype(str).eq(row_id), "Parent Row ID"] = new_parent_row_id
    target_subtree = _subtree_ids(remainder, new_parent_row_id)
    insert_at = max(int(remainder.index[remainder["Row ID"].astype(str).eq(rid)][0]) for rid in target_subtree) + 1
    combined = pd.concat([remainder.iloc[:insert_at], block, remainder.iloc[insert_at:]], ignore_index=True)
    return refresh_schedule_calculations(ensure_schedule_hierarchy(combined))


def outdent_row(schedule_df: pd.DataFrame, row_id: str) -> pd.DataFrame:
    """Move a subtree one level up, immediately after its current parent subtree."""
    df = ensure_schedule_hierarchy(schedule_df)
    match = df[df["Row ID"].astype(str).eq(str(row_id))]
    if match.empty:
        return df
    parent_id = clean_text(match.iloc[0]["Parent Row ID"])
    if not parent_id:
        return df
    parent_match = df[df["Row ID"].astype(str).eq(parent_id)]
    grandparent_id = clean_text(parent_match.iloc[0]["Parent Row ID"]) if not parent_match.empty else ""
    subtree_ids = _subtree_ids(df, str(row_id))
    block = df[df["Row ID"].astype(str).isin(subtree_ids)].copy()
    remainder = df[~df["Row ID"].astype(str).isin(subtree_ids)].copy().reset_index(drop=True)
    block.loc[block["Row ID"].astype(str).eq(str(row_id)), "Parent Row ID"] = grandparent_id
    parent_subtree = _subtree_ids(remainder, parent_id)
    insert_at = max(int(remainder.index[remainder["Row ID"].astype(str).eq(rid)][0]) for rid in parent_subtree) + 1
    combined = pd.concat([remainder.iloc[:insert_at], block, remainder.iloc[insert_at:]], ignore_index=True)
    return refresh_schedule_calculations(ensure_schedule_hierarchy(combined))


def nest_group_into(schedule_df: pd.DataFrame, child_group_id: str, target_group_id: str) -> pd.DataFrame:
    """Place one complete parent block beneath another as a nested subassembly.

    The child root remains bold through Is Assembly Root, receives the next
    available alphabetic nested code, and keeps its complete subtree.
    """
    df = ensure_schedule_hierarchy(schedule_df)
    child = df[df["Group ID"].astype(str).eq(str(child_group_id))]
    target = df[df["Group ID"].astype(str).eq(str(target_group_id))]
    if child.empty or target.empty:
        return df
    return reparent_row(df, str(child.iloc[0]["Row ID"]), str(target.iloc[0]["Row ID"]))


def make_custom_schedule_row(item: str = "CUSTOM", description: str = "Engineer-added project item") -> Dict[str, Any]:
    """Create a project-only row with visibly unresolved requirement fields."""
    row: Dict[str, Any] = {col: "" for col in SCHEDULE_COLUMNS}
    row.update({
        "Nested Code": "Parent",
        "Is Assembly Root": False,
        "Item": clean_text(item) or "CUSTOM",
        "Description": clean_text(description) or "Engineer-added project item",
        "Count Category": component_count_category(item, description),
        "Qty": 1,
        "Source Quote Qty": 1,
        "Match Type": "Engineer Custom Item",
        "Requirements Source": "Project-only custom row",
        "Requirements Status": "Custom / requirements required",
        "Bold": False,
        "Italic": False,
        "Underline": False,
        "Highlight": False,
        "Review Added": False,
    })
    for column in REQUIREMENT_COLUMNS:
        row[column] = "-"
    return row


def _review_match_key(item: Any, description: Any) -> Tuple[str, str]:
    """Stable exact comparison key for Review-to-draft count checks."""
    return norm(item), clean_text(description).upper()


def review_item_count_snapshot(
    schedule_df: pd.DataFrame,
    review_df: pd.DataFrame,
) -> pd.DataFrame:
    """Add exact draft-count and missing-quantity guidance to Review rows.

    Demand is aggregated for repeated Review lines with the same exact item and
    description. This prevents adding the same missing quantity twice when an
    engineer selects several duplicate Review lines at once.
    """
    review = review_df.copy() if review_df is not None else pd.DataFrame()
    if review.empty:
        for column in [
            "Exact Item Rows in Draft", "Exact Description Rows in Draft",
            "Quote Demand", "Suggested Rows To Add", "Draft Count Check",
        ]:
            review[column] = pd.Series(dtype="object")
        return review

    schedule = ensure_schedule_hierarchy(schedule_df)
    item_counts: Dict[str, int] = {}
    variant_counts: Dict[Tuple[str, str], int] = {}
    for _, row in schedule.iterrows():
        item_key, variant_description = _review_match_key(
            row.get("Item"), row.get("Description")
        )
        if not item_key:
            continue
        item_counts[item_key] = item_counts.get(item_key, 0) + 1
        variant_key = (item_key, variant_description)
        variant_counts[variant_key] = variant_counts.get(variant_key, 0) + 1

    demand: Dict[Tuple[str, str], int] = {}
    for _, row in review.iterrows():
        key = _review_match_key(row.get("Item"), row.get("Description"))
        if not key[0]:
            continue
        qty = safe_float(row.get("Qty"))
        requested = int(round(qty)) if qty is not None and is_whole_number(qty) and qty >= 1 else 1
        demand[key] = demand.get(key, 0) + requested

    exact_item_rows: List[int] = []
    exact_variant_rows: List[int] = []
    quote_demand: List[int] = []
    suggested: List[int] = []
    checks: List[str] = []
    for _, row in review.iterrows():
        key = _review_match_key(row.get("Item"), row.get("Description"))
        code_count = item_counts.get(key[0], 0)
        variant_count = variant_counts.get(key, 0)
        requested = demand.get(key, 0)
        missing = max(0, requested - variant_count)
        exact_item_rows.append(code_count)
        exact_variant_rows.append(variant_count)
        quote_demand.append(requested)
        suggested.append(missing)
        if missing:
            checks.append(f"Add {missing} missing row{'s' if missing != 1 else ''}")
        elif requested:
            checks.append("Exact item and description already represented")
        else:
            checks.append("No usable item code")

    review["Exact Item Rows in Draft"] = exact_item_rows
    review["Exact Description Rows in Draft"] = exact_variant_rows
    review["Quote Demand"] = quote_demand
    review["Suggested Rows To Add"] = suggested
    review["Draft Count Check"] = checks
    return review


def schedule_rows_for_review_items(
    schedule_df: pd.DataFrame,
    selected_review_df: pd.DataFrame,
    all_review_df: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    """Create only missing exact item/description quantities from Review."""
    selected = selected_review_df.copy() if selected_review_df is not None else pd.DataFrame()
    all_review = all_review_df if all_review_df is not None else selected
    if selected.empty:
        return pd.DataFrame(columns=SCHEDULE_COLUMNS), []

    snapshot = review_item_count_snapshot(schedule_df, all_review)
    snapshot_by_key: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for _, row in snapshot.iterrows():
        key = _review_match_key(row.get("Item"), row.get("Description"))
        if key[0] and key not in snapshot_by_key:
            snapshot_by_key[key] = row.to_dict()

    rows: List[Dict[str, Any]] = []
    report: List[Dict[str, Any]] = []
    processed: set[Tuple[str, str]] = set()
    for _, selected_row in selected.iterrows():
        key = _review_match_key(selected_row.get("Item"), selected_row.get("Description"))
        if not key[0] or key in processed:
            continue
        processed.add(key)
        count_row = snapshot_by_key.get(key, selected_row.to_dict())
        missing = int(safe_float(count_row.get("Suggested Rows To Add")) or 0)
        before = int(safe_float(count_row.get("Exact Description Rows in Draft")) or 0)
        for _ in range(missing):
            new_row = make_custom_schedule_row(
                selected_row.get("Item", "CUSTOM"),
                selected_row.get("Description", "Review item added by engineer"),
            )
            new_row.update({
                "Group ID": f"REVIEW{len(rows) + 1:04d}",
                "Qty": 1,
                "Source Quote Qty": selected_row.get("Qty", 1),
                "Source Quote Item": clean_text(selected_row.get("Item")),
                "Match Type": "Engineer Added From Review",
                "Requirements Source": "Review item added by engineer",
                "Requirements Status": "Review-added / requirements required",
                "Review Added": True,
            })
            rows.append(new_row)
        report.append({
            "Item": clean_text(selected_row.get("Item")),
            "Description": clean_text(selected_row.get("Description")),
            "Before": before,
            "Added": missing,
            "After": before + missing,
        })
    return pd.DataFrame(rows, columns=SCHEDULE_COLUMNS), report


def replace_row_from_master(
    schedule_df: pd.DataFrame,
    row_id: str,
    master: MasterList,
    excel_row: int,
) -> pd.DataFrame:
    """Replace selected row content/requirements while preserving its tree position."""
    df = ensure_schedule_hierarchy(schedule_df)
    selected = next((r for r in master.rows if r.excel_row == int(excel_row)), None)
    mask = df["Row ID"].astype(str).eq(str(row_id))
    if selected is None or not mask.any():
        return df
    replacement = schedule_row_for_master_excel_row(master, int(excel_row))
    preserved = {
        column: df.loc[mask, column].iloc[0]
        for column in [
            "Order", "Group ID", "Parent Group ID", "Row ID", "Parent Row ID",
            "Depth", "Nested Code", "Instance", "Source Quote Qty", "Is Assembly Root",
        ]
    }
    for column in SCHEDULE_COLUMNS:
        if column in preserved:
            df.loc[mask, column] = preserved[column]
        elif column in replacement:
            df.loc[mask, column] = replacement[column]
    return refresh_schedule_calculations(refresh_requirement_status(df, master))


def apply_row_format(
    schedule_df: pd.DataFrame,
    row_id: str,
    bold: Optional[bool] = None,
    italic: Optional[bool] = None,
    underline: Optional[bool] = None,
    highlight: Optional[bool] = None,
) -> pd.DataFrame:
    df = ensure_schedule_hierarchy(schedule_df)
    mask = df["Row ID"].astype(str).eq(str(row_id))
    if not mask.any():
        return df
    for column, value in {
        "Bold": bold,
        "Italic": italic,
        "Underline": underline,
        "Highlight": highlight,
    }.items():
        if value is not None:
            df.loc[mask, column] = bool(value)
    return df


def apply_rows_format(
    schedule_df: pd.DataFrame,
    row_ids: Iterable[str],
    bold: Optional[bool] = None,
    italic: Optional[bool] = None,
    underline: Optional[bool] = None,
    highlight: Optional[bool] = None,
) -> pd.DataFrame:
    """Apply formatting to exactly the selected rows, without altering hierarchy."""
    df = ensure_schedule_hierarchy(schedule_df)
    selected = {clean_text(row_id) for row_id in row_ids if clean_text(row_id)}
    mask = df["Row ID"].astype(str).isin(selected)
    if not mask.any():
        return df
    for column, value in {
        "Bold": bold,
        "Italic": italic,
        "Underline": underline,
        "Highlight": highlight,
    }.items():
        if value is not None:
            df.loc[mask, column] = bool(value)
    return df


def _requirements_equal(left: Any, right: Any) -> bool:
    left_num = safe_float(left)
    right_num = safe_float(right)
    if left_num is not None and right_num is not None:
        return abs(left_num - right_num) < 1e-9
    return clean_text(left).upper() == clean_text(right).upper()


def refresh_requirement_status(schedule_df: pd.DataFrame, master: MasterList) -> pd.DataFrame:
    """Mark each Master-derived row as inherited or visibly overridden."""
    df = ensure_schedule_hierarchy(schedule_df)
    by_excel_row = {row.excel_row: row for row in master.rows}
    attr_map = {
        "HP": "hp", "Phase": "phase", "Volts": "volts", "Amps": "amps",
        "C.B.": "cb", "Air Port": "air_port", "Cold Water": "cold_water",
        "Hot Water": "hot_water", "Reclaim Water": "reclaim_water", "Gas BTUH": "gas_btuh",
    }
    for idx, record in df.iterrows():
        excel_row = safe_float(record.get("Master Excel Row"))
        master_row = by_excel_row.get(int(excel_row)) if excel_row is not None else None
        if master_row is None:
            if clean_text(record.get("Requirements Source")).startswith("QuickBooks "):
                df.at[idx, "Requirements Status"] = (
                    "QuickBooks fallback · engineering requirements needed"
                )
                continue
            if not clean_text(record.get("Requirements Source")):
                df.at[idx, "Requirements Source"] = "Project-only custom row"
            df.at[idx, "Requirements Status"] = "Custom / requirements required"
            continue
        df.at[idx, "Requirements Source"] = f"Master row {master_row.excel_row} · {master_row.part_number}"
        inherited = all(
            _requirements_equal(record.get(column), getattr(master_row, attr_name))
            for column, attr_name in attr_map.items()
        )
        df.at[idx, "Requirements Status"] = (
            "Inherited from Master" if inherited else "Engineer override"
        )
    return df


def reset_row_requirements_from_master(
    schedule_df: pd.DataFrame,
    row_id: str,
    master: MasterList,
) -> pd.DataFrame:
    """Restore only requirement fields for one selected row from its exact source row."""
    df = ensure_schedule_hierarchy(schedule_df)
    mask = df["Row ID"].astype(str).eq(str(row_id))
    if not mask.any():
        return df
    excel_row = safe_float(df.loc[mask, "Master Excel Row"].iloc[0])
    selected = next((r for r in master.rows if excel_row is not None and r.excel_row == int(excel_row)), None)
    if selected is None:
        return df
    attr_map = {
        "HP": "hp", "Phase": "phase", "Volts": "volts", "Amps": "amps",
        "C.B.": "cb", "Air Port": "air_port", "Cold Water": "cold_water",
        "Hot Water": "hot_water", "Reclaim Water": "reclaim_water", "Gas BTUH": "gas_btuh",
    }
    for column, attr_name in attr_map.items():
        df.loc[mask, column] = getattr(selected, attr_name)
    return refresh_schedule_calculations(refresh_requirement_status(df, master))


def apply_component_counts(schedule_df: pd.DataFrame) -> pd.DataFrame:
    """Fill # using an engineering component category sequence.

    Different motor ratings share the Motor sequence. All other parts default
    to their exact canonical item unless a Count Category is assigned.
    Quote quantity is ignored because equipment groups are already repeated.
    """
    if schedule_df is None or schedule_df.empty:
        return schedule_df
    df = schedule_df.copy()
    if "#" not in df.columns:
        insert_at = list(df.columns).index("Description") + 1 if "Description" in df.columns else len(df.columns)
        df.insert(insert_at, "#", "")
    counts: Dict[str, int] = {}
    values: List[Any] = []
    for _, row in df.iterrows():
        item_key = norm(row.get("Item"))
        count_key = clean_text(row.get("Count Category")) or component_count_category(
            row.get("Item"), row.get("Description")
        )
        if not item_key or not count_key:
            values.append("")
            continue
        counts[count_key] = counts.get(count_key, 0) + 1
        values.append(counts[count_key])
        if "Count Category" in df.columns:
            df.at[row.name, "Count Category"] = count_key
    df["#"] = values
    return df[[c for c in SCHEDULE_COLUMNS if c in df.columns]]


def refresh_schedule_calculations(schedule_df: pd.DataFrame) -> pd.DataFrame:
    """Refresh app-side calculated preview values before export."""
    if schedule_df is None or schedule_df.empty:
        return schedule_df
    df = apply_component_counts(ensure_schedule_hierarchy(schedule_df))
    # Extended-amps meaning is pending engineering confirmation. Keep the
    # preview field blank so the pilot cannot silently double-count repeated
    # groups.
    if "Ext. Amps" in df.columns:
        df["Ext. Amps"] = ""
    return df


def build_summary(meta: Dict[str, Any], schedule_df: pd.DataFrame, review_df: pd.DataFrame) -> pd.DataFrame:
    total_amps = 0.0
    volts: Dict[str, float] = {}
    if not schedule_df.empty:
        for _, row in schedule_df.iterrows():
            amps = amp_value(row.get("Ext. Amps", row.get("Amps")))
            total_amps += amps
            volt_key = normalize_voltage(row.get("Volts")) or "No Voltage"
            volts[volt_key] = volts.get(volt_key, 0.0) + amps

    rows = [
        ["Customer", meta.get("customer", "")],
        ["Invoice Number", meta.get("invoice_number", "")],
        ["Invoice Date", meta.get("invoice_date", "")],
        ["Country", meta.get("country", "")],
        ["Generated Schedule Rows", len(schedule_df)],
        ["Review Items", len(review_df)],
        ["Total Amps", round(total_amps, 2)],
        ["", ""],
        ["Voltage", "Total Amps"],
    ]
    for voltage, amps in sorted(volts.items()):
        rows.append([voltage, round(amps, 2)])
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def write_output_workbook(
    meta: Dict[str, Any],
    quote_df: pd.DataFrame,
    schedule_df: pd.DataFrame,
    review_df: pd.DataFrame,
    rules: Dict[str, pd.DataFrame],
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    schedule_df = refresh_schedule_calculations(schedule_df if schedule_df is not None else pd.DataFrame(columns=SCHEDULE_COLUMNS))
    title = f"{(meta.get('customer') or 'CUSTOMER').upper()} - ELECTRICAL SCHEDULE"

    ws2 = wb.create_sheet("Electrical Schedule")
    write_electrical_schedule_sheet(ws2, schedule_df, title)
    total_row = getattr(ws2, "_avw_total_row", ws2.max_row)
    first_data_row = getattr(ws2, "_avw_first_data_row", 7)
    last_data_row = getattr(ws2, "_avw_last_data_row", max(first_data_row, total_row - 1))
    volts_col = getattr(ws2, "_avw_volts_col", 9)
    amps_col = getattr(ws2, "_avw_amps_col", 10)

    write_summary_sheet(ws, meta, schedule_df, review_df, title, total_row, first_data_row, last_data_row, volts_col, amps_col)

    ws3 = wb.create_sheet("Quote Extract")
    write_df(ws3, quote_df if quote_df is not None and not quote_df.empty else pd.DataFrame(columns=QUOTE_COLUMNS))

    ws4 = wb.create_sheet("Review Items")
    write_df(ws4, review_df if review_df is not None and not review_df.empty else pd.DataFrame(columns=REVIEW_COLUMNS))

    for sheet in wb.worksheets:
        format_sheet(sheet)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def write_summary_sheet(
    ws,
    meta: Dict[str, Any],
    schedule_df: pd.DataFrame,
    review_df: pd.DataFrame,
    title: str,
    total_row: int,
    first_data_row: int,
    last_data_row: int,
    volts_col: int,
    amps_col: int,
) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(bold=True, size=15)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    rows = [
        ("Customer", meta.get("customer", "")),
        ("Invoice Number", meta.get("invoice_number", "")),
        ("Invoice Date", meta.get("invoice_date", "")),
        ("Country", meta.get("country", "")),
        ("Generated Schedule Rows", len(schedule_df) if schedule_df is not None else 0),
        ("Review Items", len(review_df) if review_df is not None else 0),
        ("Total Amps", f"='Electrical Schedule'!{get_column_letter(amps_col)}{total_row}"),
        ("Extended Amps", "Pending engineering definition; excluded from this pilot calculation."),
        (
            "Engineer Requirement Overrides",
            int(schedule_df.get("Requirements Status", pd.Series(dtype=str)).astype(str).eq("Engineer override").sum()),
        ),
        (
            "Custom Requirement Rows",
            int(schedule_df.get("Requirements Status", pd.Series(dtype=str)).astype(str).str.startswith("Custom").sum()),
        ),
    ]
    ws.cell(4, 1).value = "Metric"
    ws.cell(4, 2).value = "Value"
    for idx, (metric, value) in enumerate(rows, start=5):
        ws.cell(idx, 1).value = metric
        ws.cell(idx, 2).value = value

    voltage_values: List[Any] = []
    seen = set()
    if schedule_df is not None and not schedule_df.empty and "Volts" in schedule_df.columns:
        for value in schedule_df["Volts"].tolist():
            text = clean_text(value)
            if not text or text == "-":
                continue
            key = text.upper()
            if key not in seen:
                seen.add(key)
                voltage_values.append(value)

    start = 5 + len(rows) + 1
    ws.cell(start, 1).value = "Voltage"
    ws.cell(start, 2).value = "Total Amps"
    volts_letter = get_column_letter(volts_col)
    amps_letter = get_column_letter(amps_col)
    volts_range = f"'Electrical Schedule'!{volts_letter}${first_data_row}:{volts_letter}${last_data_row}"
    amps_range = f"'Electrical Schedule'!{amps_letter}${first_data_row}:{amps_letter}${last_data_row}"
    for idx, voltage in enumerate(voltage_values, start=start + 1):
        ws.cell(idx, 1).value = voltage
        ws.cell(idx, 2).value = f"=SUMIF({volts_range},A{idx},{amps_range})"


def write_electrical_schedule_sheet(ws, schedule_df: pd.DataFrame, title: str) -> None:
    schedule_df = refresh_schedule_calculations(schedule_df if schedule_df is not None else pd.DataFrame(columns=SCHEDULE_COLUMNS))
    navy = "1F4E78"
    light = "D9EAF7"
    white = "FFFFFF"
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:Q2")
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(bold=True, size=16, color=white)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=navy)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F5")
    ws["A3"] = "CAR WASH EQUIPMENT LIST"
    ws.merge_cells("G3:Q4")
    ws["G3"] = "EQUIPMENT REQUIREMENTS"
    ws.merge_cells("G5:K5")
    ws["G5"] = "ELECTRICAL"
    ws["L5"] = "AIR"
    ws.merge_cells("M5:P5")
    ws["M5"] = "WATER"
    ws["Q5"] = "GAS"

    headers = {
        2: "PROJECT ITEM #", 4: "PART #", 5: "#", 6: "DESCRIPTION",
        7: "HP", 8: "PHASE", 9: "VOLTS", 10: "AMPS", 11: "C.B.",
        12: "PORT", 13: "COLD", 14: "HOT", 15: "RECLAIM", 16: "", 17: "(BTUH)",
    }
    for col, value in headers.items():
        ws.cell(6, col).value = value

    for r in range(3, 7):
        for c in range(1, 18):
            cell = ws.cell(r, c)
            cell.fill = PatternFill("solid", fgColor=light)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    first_data_row = 7
    engineering_map = [
        (4, "Item"), (5, "#"), (6, "Description"), (7, "HP"), (8, "Phase"),
        (9, "Volts"), (10, "Amps"), (11, "C.B."), (12, "Air Port"),
        (13, "Cold Water"), (14, "Hot Water"), (15, "Reclaim Water"), (17, "Gas BTUH"),
    ]
    for offset, (_, row) in enumerate(schedule_df.iterrows()):
        r_idx = first_data_row + offset
        ws.cell(r_idx, 1).value = offset + 1
        ws.cell(r_idx, 2).value = row.get("Order", "")
        nested = clean_text(row.get("Nested Code"))
        ws.cell(r_idx, 3).value = "" if nested.lower() == "parent" else nested
        for c_idx, key in engineering_map:
            value = row.get(key, "")
            if pd.isna(value):
                value = ""
            ws.cell(r_idx, c_idx).value = value
        source = clean_text(row.get("Requirements Source")) or "No Master source"
        status = clean_text(row.get("Requirements Status")) or "Unknown"
        ws.cell(r_idx, 4).comment = Comment(
            f"Requirements source: {source}\nStatus: {status}",
            "AVW Schedule Generator",
        )
        is_root = bool(row.get("Is Assembly Root")) or nested.lower() == "parent"
        row_bold = is_root or enabled_bool(row.get("Bold"))
        row_italic = enabled_bool(row.get("Italic"))
        row_underline = "single" if enabled_bool(row.get("Underline")) else None
        row_highlight = enabled_bool(row.get("Highlight"))
        review_added = enabled_bool(row.get("Review Added"))
        for c in range(1, 18):
            ws.cell(r_idx, c).border = border
            ws.cell(r_idx, c).alignment = Alignment(vertical="top", wrap_text=True)
            if review_added:
                ws.cell(r_idx, c).fill = PatternFill("solid", fgColor="F4CCCC")
            elif row_highlight:
                ws.cell(r_idx, c).fill = PatternFill("solid", fgColor="FFF2B2")
            ws.cell(r_idx, c).font = Font(
                bold=row_bold,
                italic=row_italic,
                underline=row_underline,
            )
        description_text = clean_text(row.get("Description"))
        estimated_lines = max(1, math.ceil(len(description_text) / 52))
        ws.row_dimensions[r_idx].height = max(18, min(72, estimated_lines * 14))

    last_data_row = first_data_row + len(schedule_df) - 1
    total_row = max(first_data_row, last_data_row + 1)
    if schedule_df.empty:
        last_data_row = first_data_row - 1
        total_row = first_data_row

    ws.cell(total_row, 6).value = "TOTAL AMPS"
    ws.cell(total_row, 6).font = Font(bold=True)
    ws.cell(total_row, 10).value = f"=SUM(J{first_data_row}:J{last_data_row})" if last_data_row >= first_data_row else 0
    ws.cell(total_row, 10).font = Font(bold=True)
    for c in range(1, 18):
        ws.cell(total_row, c).border = border

    widths = {"A": 6, "B": 13, "C": 9, "D": 24, "E": 7, "F": 52,
              "G": 9, "H": 9, "I": 10, "J": 10, "K": 10, "L": 11,
              "M": 11, "N": 11, "O": 12, "P": 8, "Q": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 23
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 30
    ws.freeze_panes = "A7"
    ws.print_title_rows = "1:6"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:Q{total_row}"

    ws._avw_first_data_row = first_data_row
    ws._avw_last_data_row = last_data_row
    ws._avw_total_row = total_row
    ws._avw_volts_col = 9
    ws._avw_amps_col = 10

def write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    if df is None or df.empty:
        return
    for c_idx, col in enumerate(df.columns, start=start_col):
        ws.cell(start_row, c_idx).value = col
    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for c_idx, col in enumerate(df.columns, start=start_col):
            value = row[col]
            if pd.isna(value):
                value = ""
            ws.cell(r_idx, c_idx).value = value


def format_sheet(ws) -> None:
    # The engineering schedule has a fixed AVW-style layout prepared by its
    # dedicated writer; keep its alignments, widths, merges, and print setup.
    if ws.title == "Electrical Schedule":
        return

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    # format likely header rows
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [ws.cell(row_idx, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        if values and all(v is not None for v in values[: min(len(values), 2)]):
            first = str(values[0]).strip().lower()
            if first in {"order", "metric", "topic", "enabled"} or row_idx == 1:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row_idx, c).font = Font(bold=True)
                    ws.cell(row_idx, c).fill = header_fill
    if ws.title != "Electrical Schedule" and ws.max_row > 1 and ws.max_column > 1:
        ws.freeze_panes = "A2"
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for cell in ws[letter]:
            text = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, min(len(text), 60))
        ws.column_dimensions[letter].width = min(max_len + 2, 45)


def process_files(
    quote_pdf: Any,
    master_excel: Any,
    quickbooks_catalog: Optional[QuickBooksCatalog] = None,
    alias_rules: Optional[Any] = None,
    replacement_rules: Optional[Any] = None,
    special_rules: Optional[Any] = None,
    ignore_rules: Optional[Any] = None,
) -> Dict[str, Any]:
    meta, quote_lines, raw_text = parse_quote_pdf(quote_pdf)
    quote_df = quote_lines_to_df(quote_lines)
    master = MasterList.from_excel(master_excel)

    alias_df = read_rules_csv(alias_rules)
    replacement_df = read_rules_csv(replacement_rules)
    special_df = read_rules_csv(special_rules)
    ignore_df = read_rules_csv(ignore_rules)

    schedule_df, review_df = generate_schedule(
        quote_lines,
        master,
        quickbooks=quickbooks_catalog,
        alias_df=alias_df,
        replacement_df=replacement_df,
        special_df=special_df,
        ignore_df=ignore_df,
    )
    output = write_output_workbook(
        meta,
        quote_df,
        schedule_df,
        review_df,
        rules={
            "Alias Rules": alias_df,
            "Replacement Rules": replacement_df,
            "Special Quantity Rules": special_df,
            "Ignore Rules": ignore_df,
        },
    )
    return {
        "meta": meta,
        "quote_df": quote_df,
        "schedule_df": schedule_df,
        "review_df": review_df,
        "output_xlsx": output,
        "master": master,
        "parent_catalog_df": build_parent_catalog_df(master),
        "component_catalog_df": build_component_catalog_df(master),
        "quickbooks": quickbooks_catalog,
        "quickbooks_catalog_df": build_quickbooks_catalog_df(quickbooks_catalog),
        "rules": {
            "Alias Rules": alias_df,
            "Replacement Rules": replacement_df,
            "Special Quantity Rules": special_df,
            "Ignore Rules": ignore_df,
        },
        "master_row_count": len(master.rows),
        "main_equipment_reference_count": len(master.main_equipment_reference),
        "quickbooks_item_count": len(quickbooks_catalog.items) if quickbooks_catalog else 0,
    }

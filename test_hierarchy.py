"""Focused regression checks for editable schedule hierarchy and Master inheritance."""

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from avw_schedule.processor import (
    MasterList,
    MasterRow,
    QuickBooksCatalog,
    QuickBooksItem,
    QuoteLine,
    SCHEDULE_COLUMNS,
    apply_row_format,
    apply_rows_format,
    delete_row_subtrees,
    ensure_schedule_hierarchy,
    generate_schedule,
    insert_schedule_rows,
    make_custom_schedule_row,
    move_row_subtree,
    move_row_subtrees,
    outdent_row,
    refresh_requirement_status,
    reparent_row,
    review_item_count_snapshot,
    reset_row_requirements_from_master,
    schedule_rows_for_review_items,
    schedule_row_for_quickbooks_excel_row,
    subtree_descendant_counts,
    write_output_workbook,
)


def custom(item: str, code: str, group: str = "G001") -> dict:
    row = make_custom_schedule_row(item, item)
    row["Group ID"] = group
    row["Nested Code"] = code
    row["Is Assembly Root"] = code == "Parent"
    return row


def assert_nested_code_reflow() -> None:
    original = pd.DataFrame([
        custom("ROOT", "Parent"),
        custom("FIRST", "A"),
        custom("GRANDCHILD", "AA"),
        custom("SECOND", "B"),
    ], columns=SCHEDULE_COLUMNS)
    schedule = ensure_schedule_hierarchy(original)
    ids = schedule.set_index("Item")["Row ID"].to_dict()
    assert schedule.set_index("Item")["Nested Code"].to_dict() == {
        "ROOT": "Parent", "FIRST": "A", "GRANDCHILD": "AA", "SECOND": "B",
    }

    inserted = pd.DataFrame([custom("NEW", "Parent", "G999")], columns=SCHEDULE_COLUMNS)
    schedule = insert_schedule_rows(schedule, inserted, mode="before", target_row_id=ids["FIRST"])
    codes = schedule.set_index("Item")["Nested Code"].to_dict()
    assert codes["NEW"] == "A"
    assert codes["FIRST"] == "B"
    assert codes["GRANDCHILD"] == "BA"
    assert codes["SECOND"] == "C"

    ids = schedule.set_index("Item")["Row ID"].to_dict()
    schedule = move_row_subtree(schedule, ids["SECOND"], -1)
    codes = schedule.set_index("Item")["Nested Code"].to_dict()
    assert codes["SECOND"] == "B" and codes["FIRST"] == "C" and codes["GRANDCHILD"] == "CA"

    schedule = reparent_row(schedule, ids["SECOND"], ids["FIRST"])
    codes = schedule.set_index("Item")["Nested Code"].to_dict()
    assert codes["SECOND"] == "BB"
    schedule = outdent_row(schedule, ids["SECOND"])
    codes = schedule.set_index("Item")["Nested Code"].to_dict()
    assert codes["SECOND"] == "C"


def assert_master_inheritance_and_export_style() -> None:
    master_row = MasterRow(
        excel_row=10, group_id=1, nested_code="", part_number="M-20",
        description="Motor", hp=20, phase=3, volts=460, amps=24,
        cb=30, air_port="-", cold_water="-", hot_water="-",
        reclaim_water="-", gas_btuh="-", is_parent=True,
    )
    master = MasterList([master_row], [])
    schedule = pd.DataFrame([
        master_row.to_schedule_row("1", "G001", 1, "M-20", "Motor", "Exact Parent")
    ], columns=SCHEDULE_COLUMNS)
    schedule = ensure_schedule_hierarchy(schedule)
    row_id = str(schedule.iloc[0]["Row ID"])
    schedule = refresh_requirement_status(schedule, master)
    assert schedule.iloc[0]["Requirements Status"] == "Inherited from Master"
    schedule.loc[0, "Volts"] = 575
    schedule = refresh_requirement_status(schedule, master)
    assert schedule.iloc[0]["Requirements Status"] == "Engineer override"
    schedule = reset_row_requirements_from_master(schedule, row_id, master)
    assert schedule.iloc[0]["Volts"] == 460
    assert schedule.iloc[0]["Requirements Status"] == "Inherited from Master"

    schedule = apply_row_format(
        schedule, row_id, bold=True, italic=True, underline=True, highlight=True
    )
    output = write_output_workbook(
        {"customer": "Test"}, pd.DataFrame(), schedule, pd.DataFrame(), rules={}
    )
    workbook = load_workbook(BytesIO(output.getvalue()), data_only=False)
    font = workbook["Electrical Schedule"]["D7"].font
    assert font.bold and font.italic and font.underline == "single"
    assert workbook["Electrical Schedule"]["D7"].fill.fgColor.rgb == "00FFF2B2"


def assert_batch_selection_actions_keep_subtrees_safe() -> None:
    original = pd.DataFrame([
        custom("ROOT-1", "Parent", "G001"),
        custom("ROOT-1-A", "A", "G001"),
        custom("ROOT-1-AA", "AA", "G001"),
        custom("ROOT-2", "Parent", "G002"),
        custom("ROOT-2-A", "A", "G002"),
        custom("ROOT-3", "Parent", "G003"),
    ], columns=SCHEDULE_COLUMNS)
    schedule = ensure_schedule_hierarchy(original)
    ids = schedule.set_index("Item")["Row ID"].to_dict()
    descendant_counts = subtree_descendant_counts(schedule)
    assert descendant_counts[ids["ROOT-1"]] == 2
    assert descendant_counts[ids["ROOT-1-A"]] == 1
    assert descendant_counts[ids["ROOT-3"]] == 0

    # Moving two roots uses schedule order, includes descendants, and needs one action.
    schedule = move_row_subtrees(
        schedule,
        [ids["ROOT-3"], ids["ROOT-1"]],
        target_row_id=ids["ROOT-2"],
        mode="after",
    )
    roots = schedule[schedule["Depth"].eq(0)]["Item"].tolist()
    assert roots == ["ROOT-2", "ROOT-1", "ROOT-3"]
    assert schedule["Item"].tolist() == [
        "ROOT-2", "ROOT-2-A", "ROOT-1", "ROOT-1-A", "ROOT-1-AA", "ROOT-3"
    ]
    codes = schedule.set_index("Item")["Nested Code"].to_dict()
    assert codes["ROOT-1-A"] == "A" and codes["ROOT-1-AA"] == "AA"

    # Formatting affects exactly the selected rows and survives Excel export.
    schedule = apply_rows_format(
        schedule,
        [ids["ROOT-2-A"], ids["ROOT-3"]],
        bold=True,
        highlight=True,
    )
    formatted = schedule.set_index("Item")
    assert bool(formatted.at["ROOT-2-A", "Bold"])
    assert bool(formatted.at["ROOT-3", "Highlight"])
    assert not bool(formatted.at["ROOT-1-A", "Highlight"])

    # Selecting a parent and its child deletes the subtree once, not twice.
    schedule = delete_row_subtrees(schedule, [ids["ROOT-1"], ids["ROOT-1-A"]])
    assert "ROOT-1" not in schedule["Item"].tolist()
    assert "ROOT-1-A" not in schedule["Item"].tolist()
    assert "ROOT-1-AA" not in schedule["Item"].tolist()


def assert_total_prefixed_component_does_not_end_master_list() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MASTER LIST"

    sheet["B7"] = 1
    sheet["D7"] = "SOB-RO-10GPM"
    sheet["F7"] = "REVERSE OSMOSIS UNIT"
    sheet["D7"].font = Font(bold=True)

    sheet["C8"] = "A"
    sheet["D8"] = "TDS"
    sheet["F8"] = "TOTAL DISSOLVED SOLIDS MONITOR"

    sheet["B9"] = 2
    sheet["D9"] = "ITEM-AFTER-TDS"
    sheet["F9"] = "VALID ITEM AFTER TOTAL-PREFIXED COMPONENT"
    sheet["D9"].font = Font(bold=True)

    sheet["F10"] = "TOTAL  Σ"
    sheet["B11"] = 3
    sheet["D11"] = "AFTER-FOOTER"
    sheet["F11"] = "MUST NOT LOAD"
    sheet["D11"].font = Font(bold=True)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    master = MasterList.from_excel(stream)
    loaded_items = [row.part_number for row in master.rows]
    assert loaded_items == ["SOB-RO-10GPM", "TDS", "ITEM-AFTER-TDS"]
    assert master.find_component("TDS")[0] == "single"
    assert master.find_parent("ITEM-AFTER-TDS")[0] == "single"


def assert_review_items_add_only_missing_counts_and_export_red() -> None:
    schedule = pd.DataFrame([
        custom("UNMATCHED-1", "Parent", "G001"),
        custom("OTHER", "Parent", "G002"),
    ], columns=SCHEDULE_COLUMNS)
    schedule.loc[0, "Description"] = "Exact unresolved component"
    schedule = ensure_schedule_hierarchy(schedule)
    review = pd.DataFrame([
        {
            "Order": 10, "Item": "UNMATCHED-1",
            "Description": "Exact unresolved component", "Qty": 3,
            "Issue Type": "Unmatched", "Details": "Not in Master",
        },
        {
            "Order": 11, "Item": "UNMATCHED-1",
            "Description": "Different configuration", "Qty": 1,
            "Issue Type": "Unmatched", "Details": "Not in Master",
        },
    ])
    snapshot = review_item_count_snapshot(schedule, review)
    assert snapshot.loc[0, "Exact Item Rows in Draft"] == 1
    assert snapshot.loc[0, "Exact Description Rows in Draft"] == 1
    assert snapshot.loc[0, "Suggested Rows To Add"] == 2
    assert snapshot.loc[1, "Suggested Rows To Add"] == 1

    new_rows, report = schedule_rows_for_review_items(schedule, review, review)
    assert len(new_rows) == 3
    assert new_rows["Review Added"].astype(bool).all()
    assert sum(entry["Added"] for entry in report) == 3

    first_id = str(schedule.iloc[0]["Row ID"])
    inserted = insert_schedule_rows(schedule, new_rows.iloc[:2], mode="top")
    assert inserted.iloc[0]["Item"] == "UNMATCHED-1"
    assert str(inserted.iloc[-2]["Row ID"]) == first_id

    one_review_row = insert_schedule_rows(schedule, new_rows.iloc[:1], mode="bottom")
    output = write_output_workbook(
        {"customer": "Test"}, pd.DataFrame(), one_review_row, review, rules={}
    )
    workbook = load_workbook(BytesIO(output.getvalue()), data_only=False)
    review_excel_row = 7 + len(one_review_row) - 1
    assert workbook["Electrical Schedule"].cell(review_excel_row, 4).fill.fgColor.rgb == "00F4CCCC"


def master_component(
    row: int,
    group: int,
    code: str,
    item: str,
    description: str,
    parent: bool = False,
    bold: bool = False,
) -> MasterRow:
    return MasterRow(
        excel_row=row, group_id=group, nested_code=code,
        part_number=item, description=description,
        hp="-", phase="-", volts="-", amps="-", cb="-",
        air_port="-", cold_water="-", hot_water="-",
        reclaim_water="-", gas_btuh="-",
        is_parent=parent, is_bold=bold or parent,
    )


def assert_canonical_renames_and_required_hierarchy() -> None:
    pwb = master_component(10, 1, "", "PWB1", "Pivoting wheel blaster", parent=True)
    wa_panel = master_component(11, 1, "A", "WA1-120V", "Air control panel")
    sol = master_component(12, 1, "AA", "SOL", "Solenoid")
    optional_psh = master_component(13, 1, "B", "PSH25", "Pumping station", bold=True)
    optional_motor = master_component(14, 1, "BA", "M-15", "Motor")
    optional_switch = master_component(15, 1, "BB", "LLS", "Low level switch")

    kit = master_component(20, 2, "", "WA1-SK-2018", "Wrap air assist kit", parent=True)
    kit_solenoid = master_component(21, 2, "A", "SOL", "Solenoid")
    panel = master_component(30, 3, "", "WA1-SKP-2018", "Air assist kit panel", parent=True)
    conveyor = master_component(40, 4, "", "CN1-3524", "Conveyor", parent=True)
    standalone_air = master_component(50, 5, "", "WA1-120V", "Air control panel", parent=True)
    master = MasterList([
        pwb, wa_panel, sol, optional_psh, optional_motor, optional_switch,
        kit, kit_solenoid, panel, conveyor, standalone_air,
    ], [])

    pwb_group = master.group_for_parent(pwb)
    assert [row.part_number for row in pwb_group] == ["PWB1", "WA1-120V", "SOL"]

    quote = [
        QuoteLine(1, "PWB1", "Pivoting wheel blaster", 1, 0, 0, 1, ""),
        QuoteLine(2, "WA1-SK", "Wrap air assist kit", 1, 0, 0, 1, ""),
        QuoteLine(3, "CN1", "Conveyor", 1, 0, 0, 1, ""),
        QuoteLine(4, "WA1P", "Air control panel", 1, 0, 0, 1, ""),
    ]
    schedule, review = generate_schedule(quote, master)
    assert review.empty
    assert "PSH25" not in schedule[schedule["Order"].eq("1")]["Item"].tolist()

    kit_rows = schedule[schedule["Order"].eq("2")]
    assert kit_rows["Item"].tolist() == ["WA1-SK-2018", "WA1-SKP-2018"]
    assert kit_rows["Nested Code"].tolist() == ["Parent", "A"]
    assert kit_rows.iloc[1]["Parent Row ID"] == kit_rows.iloc[0]["Row ID"]
    assert schedule[schedule["Order"].eq("3")].iloc[0]["Item"] == "CN1-3524"
    assert schedule[schedule["Order"].eq("4")].iloc[0]["Item"] == "WA1-120V"


def assert_master_first_quickbooks_fallback_and_manual_add() -> None:
    master_item = master_component(
        10, 1, "", "BOTH-1", "Engineering Master description", parent=True
    )
    ambiguous_master_a = master_component(
        11, 2, "", "AMB-MASTER", "First Master assembly", parent=True
    )
    ambiguous_master_b = master_component(
        12, 3, "", "AMB-MASTER", "Different Master assembly", parent=True
    )
    master = MasterList([master_item, ambiguous_master_a, ambiguous_master_b], [])
    quickbooks = QuickBooksCatalog([
        QuickBooksItem(4, "BOTH-1", "QuickBooks description must not override Master"),
        QuickBooksItem(5, "QB-ONLY", "QuickBooks-only standalone equipment"),
        QuickBooksItem(6, "QB-AMB", "First QuickBooks description"),
        QuickBooksItem(7, "QB-AMB", "Different QuickBooks description"),
        QuickBooksItem(8, "AMB-MASTER", "Safe exact QuickBooks fallback"),
    ])
    quote = [
        QuoteLine(1, "BOTH-1", "Quote description", 1),
        QuoteLine(2, "QB-ONLY", "Quote text must not override QuickBooks", 500),
        QuoteLine(3, "QB-AMB", "Ambiguous", 1),
        QuoteLine(4, "AMB-MASTER", "Master is ambiguous", 1),
    ]
    schedule, review = generate_schedule(quote, master, quickbooks=quickbooks)

    master_result = schedule[schedule["Order"].eq("1")].iloc[0]
    assert master_result["Description"] == "Quote description"
    assert str(master_result["Requirements Source"]).startswith("Master row 10")

    quickbooks_result = schedule[schedule["Order"].eq("2")].iloc[0]
    assert quickbooks_result["Item"] == "QB-ONLY"
    assert quickbooks_result["Description"] == "QuickBooks-only standalone equipment"
    assert quickbooks_result["Match Type"] == "QuickBooks Exact Fallback"
    assert "QuickBooks fallback" in quickbooks_result["Requirements Status"]
    assert quickbooks_result["HP"] == "-"
    assert quickbooks_result["Qty"] == 500
    assert len(schedule[schedule["Order"].eq("2")]) == 1

    assert "QB-AMB" not in schedule["Item"].tolist()
    assert review["Issue Type"].tolist() == ["Ambiguous QuickBooks Match"]

    master_ambiguous_result = schedule[
        schedule["Source Quote Item"].eq("AMB-MASTER")
    ].iloc[0]
    assert master_ambiguous_result["Description"] == "Safe exact QuickBooks fallback"
    assert master_ambiguous_result["Match Type"] == "QuickBooks Exact Fallback"

    manual = schedule_row_for_quickbooks_excel_row(quickbooks, 5, group_id="G999")
    assert manual["Item"] == "QB-ONLY"
    assert manual["Description"] == "QuickBooks-only standalone equipment"
    assert manual["Match Type"] == "Manual Add From QuickBooks"


def assert_bundled_quickbooks_catalog_parses() -> None:
    catalog_path = Path(__file__).resolve().parent / "data" / "QuickBooks_Items.xlsx"
    catalog = QuickBooksCatalog.from_excel(catalog_path)
    assert len(catalog.items) == 4233
    status, matches = catalog.find_exact("AA0A")
    assert status == "single"
    assert matches[0].description.startswith("Leg, for Applicator Arch")


if __name__ == "__main__":
    assert_nested_code_reflow()
    assert_master_inheritance_and_export_style()
    assert_batch_selection_actions_keep_subtrees_safe()
    assert_total_prefixed_component_does_not_end_master_list()
    assert_review_items_add_only_missing_counts_and_export_red()
    assert_canonical_renames_and_required_hierarchy()
    assert_master_first_quickbooks_fallback_and_manual_add()
    assert_bundled_quickbooks_catalog_parses()
    print("Hierarchy, Master/QuickBooks precedence, and export formatting checks passed.")

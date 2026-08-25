"""Run the core generator without starting Streamlit.

Usage:
    python smoke_test.py quote.pdf master.xlsx
"""

from pathlib import Path
import sys

from openpyxl import load_workbook

from avw_schedule.processor import QuickBooksCatalog, process_files


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: python smoke_test.py quote.pdf master.xlsx")
        return 2

    quote_path = Path(sys.argv[1])
    master_path = Path(sys.argv[2])
    quickbooks_path = Path(__file__).resolve().parent / "data" / "QuickBooks_Items.xlsx"
    quickbooks_catalog = QuickBooksCatalog.from_excel(quickbooks_path)
    with quote_path.open("rb") as quote_file, master_path.open("rb") as master_file:
        result = process_files(
            quote_file, master_file, quickbooks_catalog=quickbooks_catalog
        )

    workbook = load_workbook(result["output_xlsx"], data_only=False)
    expected = ["Summary", "Electrical Schedule", "Quote Extract", "Review Items"]
    assert workbook.sheetnames == expected, workbook.sheetnames
    assert len(result["quote_df"]) > 0
    assert "Unmatched" in set(result["review_df"].get("Issue Type", [])) or len(result["review_df"]) == 0

    print(f"Quote rows: {len(result['quote_df'])}")
    print(f"Schedule rows: {len(result['schedule_df'])}")
    print(f"Review rows: {len(result['review_df'])}")
    print(f"QuickBooks backend items: {result['quickbooks_item_count']}")
    print("Workbook sheets:", ", ".join(workbook.sheetnames))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
